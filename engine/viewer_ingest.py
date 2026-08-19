#!/usr/bin/env python3
"""THE VIEWER -- ingestion + OCR indexing pipeline (offline, resumable, backwards-compatible).

PDF text: PyMuPDF (pip 'pymupdf'); falls back to Poppler pdftotext.
OCR for scanned pages: RapidOCR (pip 'rapidocr-onnxruntime', no admin, bundles models);
falls back to Tesseract if RapidOCR is unavailable. Page rasterized via PyMuPDF (or pdftoppm).
OCR parallelism uses THREADS (shared engine; PyMuPDF render under a lock) -- reliable on Windows.
"""
import argparse, hashlib, json, os, re, sqlite3, subprocess, tempfile, threading, time
from concurrent.futures import ThreadPoolExecutor, as_completed
from patterns import norm_nsn   # canonical NSN normalization (A6: single source of truth)
try:
    import pymupdf as fitz  # PyMuPDF
except Exception:
    fitz = None
try:
    import numpy as _np
except Exception:
    _np = None
try:
    import ocrprep
except Exception:
    ocrprep = None
# OCR pre-processing (deskew/denoise/binarize) before every OCR call (finding #14): ocrprep.py
# already had a working, self-tested preprocessing pipeline, but nothing in this file ever called
# it -- every scanned page was OCR'd raw. Operator-toggleable (VIEWER_OCR_PREPROCESS=0 to disable)
# because the real accuracy effect on this corpus hasn't been benchmarked yet -- default on, but
# easy to turn off if it turns out to hurt more than it helps. See _ocr_preprocessed_input() below
# for why RapidOCR and Tesseract get different pipelines (preprocess_light vs. preprocess).
OCR_PREPROCESS = os.environ.get("VIEWER_OCR_PREPROCESS", "1") != "0"
try:
    import barcodes
except Exception:
    barcodes = None
# Barcode/QR read on the same page render OCR already produced (catalog §4.9): barcodes.py has had a
# fully-built, self-tested, dual-backend (pyzbar/OpenCV) detect() since it was written, but it had no
# caller anywhere in the codebase -- only its own self-test and the import-check in verifystate.py.
# Some TMs print NSNs/part numbers as barcodes; a machine-decoded value has no character-recognition
# ambiguity, so it is higher-trust provenance than OCR text (migration 0010). Operator-toggleable
# (VIEWER_BARCODE_SCAN=0 to disable) for the same reason OCR_PREPROCESS is: real-corpus benefit is
# unmeasured. Cheap regardless of the toggle -- barcodes.available() already no-ops instantly when
# neither backend is installed, so this is a no-op add-on to the existing render, never a new pass.
BARCODE_SCAN = os.environ.get("VIEWER_BARCODE_SCAN", "1") != "0"

# Dimensional-data extraction (measures.py + specparse.py + leadingspecs.py) on every page's text,
# live during ingest instead of the separate BUILD-MEASURES.bat batch pass -- pure regex on text
# already in memory, negligible cost (microseconds/page). Same opt-out convention as the two
# toggles above, offered for consistency even though this one is cheap enough that most operators
# have no reason to ever touch it.
MEASURES_SCAN = os.environ.get("VIEWER_MEASURES_SCAN", "1") != "0"
# Schematic detection (schem_overlay.py + schemgraph.py's existing netlist inference, plus a new
# keyword/caption pass for scanned pages) on every page of every newly-ingested document -- THIS
# one has a real, non-negligible cost: the vector check re-opens the PDF via PyMuPDF once per page
# (schem_overlay.schem_paths()'s own signature takes a path, not a shared handle -- same per-page
# reopen cost the existing BUILD-SCHEMGRAPH.bat batch tool already pays; not new inefficiency, just
# now running inline during the scan instead of as a separate later step). Opt-out for anyone who
# doesn't need netlist/schematic detection and wants the scan to skip that cost entirely.
SCHEMATIC_SCAN = os.environ.get("VIEWER_SCHEMATIC_SCAN", "1") != "0"
# Table extraction (tables.py's PyMuPDF find_tables(), same RPSTL/torque/PMCS/leading-particulars
# tables build_tables.py already scans for) -- another real, unwired-until-now batch tool (same
# discovery-pass finding as measures.py/schemgraph.py were): tables.db has a real schema and a real
# sidecar builder, but nothing in the live ingest path ever called it. Same per-page PDF-reopen
# cost profile as SCHEMATIC_SCAN (find_tables() needs its own PyMuPDF page handle), same opt-out
# shape for the same reason.
TABLES_SCAN = os.environ.get("VIEWER_TABLES_SCAN", "1") != "0"

OCR_CHAR_THRESHOLD = 15
NSN_RE = re.compile(r"\b\d{4}-\d{2}-\d{3}-\d{4}\b")
TM_RE  = re.compile(r"\bTM\s*[0-9][0-9A-Za-z\-]+")

_RAPID = None
OCR_DPI = 200          # render DPI for OCR (profile-tunable)
# Medium finding #24: every page used to rasterize at the fixed DPI above regardless of its
# physical size, with no output-resolution ceiling -- a large-format foldout engineering drawing
# (common in this corpus) at 200 DPI can produce a ~60+ megapixel raster handed straight to OCR/
# preprocessing with no downscale guard. _render_png() now shrinks the effective DPI for any page
# whose projected pixel count would exceed this cap; a normal-size page is completely unaffected.
OCR_MAX_MEGAPIXELS = int(os.environ.get("VIEWER_OCR_MAX_MP", "25")) * 1_000_000
# Review finding: the DPI floor (previously hardcoded at 100 with no relation to the cap above)
# could itself push a sufficiently large page's raster back OVER OCR_MAX_MEGAPIXELS, silently
# defeating the guarantee. The megapixel cap always wins; this floor only prevents a truly
# degenerate near-zero DPI on an extreme outlier page.
MIN_OCR_DPI = 20
USE_CUDA = False       # GPU OCR when True (RapidOCR + onnxruntime-gpu)
ADAPTIVE_DPI = os.environ.get("VIEWER_ADAPTIVE_DPI") == "1"   # opt-in: lower DPI on sparse pages (default OFF = no accuracy change)
_RAPID_LOCK = threading.Lock()
_FITZ_LOCK = threading.Lock()
# No timeout previously guarded either PDF rendering (PyMuPDF) or OCR inference (RapidOCR/
# Tesseract) on the default/preferred engine path -- one pathological page could stall an entire
# multi-hour ocrall() batch forever (finding #15). signal.alarm/SIGALRM isn't usable here: it can
# only preempt at a bytecode boundary or an interruptible syscall, and a single opaque native call
# (fitz.Page.get_pixmap(), an ONNX Runtime Run()) is exactly the case with neither -- that's true on
# any platform, not just Windows. A thread + .join(timeout) is the primitive that actually works:
# it gives up waiting on a real wall-clock deadline regardless of whether the target thread ever
# cooperates. The one thread that hung is leaked (never reclaimed -- Python has no safe way to kill
# a thread), but the process keeps going and the OS reclaims everything when it eventually exits.
OCR_PAGE_TIMEOUT_SECONDS = int(os.environ.get("VIEWER_OCR_PAGE_TIMEOUT", "120"))
# Separate, smaller budget for acquiring _FITZ_LOCK specifically (High-tier review finding: dual
# use of OCR_PAGE_TIMEOUT_SECONDS conflated two different things). Under heavy --workers
# contention, a worker can burn most/all of OCR_PAGE_TIMEOUT_SECONDS just queued for the
# process-wide lock, then get killed by _ocr_task's outer deadline right as it starts real work --
# indistinguishable from a genuine hang in the render/OCR call itself. A small fixed floor (not a
# fraction of the main timeout -- they govern qualitatively different things: queue-wait tolerance
# vs. total-work tolerance) lets a busy-but-healthy lock fail fast and be reported as lock
# contention, while OCR_PAGE_TIMEOUT_SECONDS remains the ceiling on the whole page.
OCR_LOCK_TIMEOUT_SECONDS = int(os.environ.get("VIEWER_OCR_LOCK_TIMEOUT", "20"))
# Review finding: the two timeouts above are independently operator-configurable via env vars with no
# cross-check. If VIEWER_OCR_PAGE_TIMEOUT is set below VIEWER_OCR_LOCK_TIMEOUT's default (e.g. a tighter
# per-page SLA), a single lock acquire could outlast the whole page's outer deadline -- the outer watchdog
# reports "timeout" and moves on while the leaked worker thread is still blocked inside acquire() for
# longer than that, defeating the split's whole point (a busy-but-healthy lock should fail fast and be
# reported as lock contention, distinct from a genuine page-level hang). Clamp, don't just document.
OCR_LOCK_TIMEOUT_SECONDS = min(OCR_LOCK_TIMEOUT_SECONDS, OCR_PAGE_TIMEOUT_SECONDS)
_DEDUP = {}                       # img_hash -> OCR text: identical pages (boilerplate) reuse text, skip re-inference
_DEDUP_LOCK = threading.Lock()
_DEDUP_STATS = {"hits": 0}
def _page_density(path, page_number):
    """Fraction of dark pixels at 50 DPI gray — a cheap blank/complexity probe (one render reused for both
    the blank-skip and the optional adaptive DPI). Returns None if PyMuPDF/numpy aren't available.
    Review finding: this calls fitz directly and used to run OUTSIDE _FITZ_LOCK even after the
    render-lock hardening below — PyMuPDF's C state isn't thread-safe, so with --workers>1 this
    probe (called at the top of every ocr_one()) could still race concurrently against another
    thread's render, exactly the class of wedge/crash _FITZ_LOCK exists to prevent. Bounded (not
    unbounded) for the same reason _render_png's acquire is bounded: a busy lock degrades this
    best-effort probe to "skip it" (default DPI, no blank-skip) instead of blocking a worker thread
    indefinitely."""
    if fitz is None or _np is None: return None
    if not _FITZ_LOCK.acquire(timeout=OCR_LOCK_TIMEOUT_SECONDS):
        return None
    try:
        doc = fitz.open(path); pix = doc[page_number-1].get_pixmap(dpi=50, colorspace=fitz.csGRAY)
        arr = _np.frombuffer(pix.samples, dtype=_np.uint8); doc.close()
        if arr.size == 0: return 0.0
        return float((arr < 110).mean())
    except Exception:
        return None
    finally:
        _FITZ_LOCK.release()
def _have_rapid():
    try:
        import rapidocr  # modern unified package (PP-OCRv5, higher accuracy)  # noqa
        return True
    except Exception:
        pass
    try:
        import rapidocr_onnxruntime  # noqa  (PP-OCRv4)
        return True
    except Exception:
        return False

def _providers():
    try:
        import onnxruntime as ort; return ort.get_available_providers()
    except Exception:
        return []

class _RapidAdapter:
    """Normalises both the modern `rapidocr` (PP-OCRv5) output and the classic
    `rapidocr_onnxruntime` (PP-OCRv4) output to the (res, elapse) shape ocr_one expects,
    where each res item is [box, text, score]."""
    def __init__(self, eng, kind): self.eng = eng; self.kind = kind
    def __call__(self, img):
        out = self.eng(img)
        if hasattr(out, "txts"):                       # modern RapidOCROutput object
            txts = list(out.txts or [])
            boxes = list(out.boxes) if getattr(out, "boxes", None) is not None else [None]*len(txts)
            scores = list(out.scores) if getattr(out, "scores", None) is not None else [1.0]*len(txts)
            res = [[boxes[i] if i < len(boxes) else None, txts[i],
                    scores[i] if i < len(scores) else 1.0] for i in range(len(txts))]
            return res, float(getattr(out, "elapse", 0) or 0)
        if isinstance(out, tuple):                     # classic (result, elapse)
            return (out[0] or []), (out[1] if len(out) > 1 else 0)
        return (out or []), 0

def _selftest(adapter):
    """Render clear synthetic text and confirm the adapter actually extracts it end-to-end.
    Guards an untested engine API: if extraction fails, we fall back to the proven path."""
    try:
        from PIL import Image, ImageDraw
        import numpy as _np
        im = Image.new("RGB", (320, 90), "white"); d = ImageDraw.Draw(im)
        d.text((12, 28), "TM 5305 24P", fill="black")
        res, _ = adapter(_np.array(im))
        txt = " ".join((r[1] or "") for r in res)
        return sum(ch.isalnum() for ch in txt) >= 3
    except Exception:
        return False

def _get_rapid(workers=1):
    global _RAPID
    if _RAPID is None:
        with _RAPID_LOCK:
            if _RAPID is None:
                _RAPID = _build_rapid(workers)
    return _RAPID

def _build_rapid(workers=1):
    prov = _providers(); gpu = "CUDAExecutionProvider" in prov
    log("OCR providers: " + (", ".join(prov) or "unknown") + (" [GPU]" if gpu else " [CPU]"))
    # Thread-count hint for the engine below: _get_rapid() builds ONE shared engine instance up
    # front (see ocr()), reused concurrently by all `workers` OCR worker THREADS. Leaving
    # onnxruntime's own intra-op thread pool at its default (-1 -> ~one thread per physical core)
    # means every concurrent session.run() call spins up its own core-sized pool, oversubscribing
    # the CPU by roughly `workers`x once workers>1 (each of `workers` concurrent calls competing
    # for all cores at once). Divide the cores across the worker threads instead so concurrent OCR
    # calls share them rather than fight over them. At workers<=1 this is a no-op (cores//1==cores,
    # same ballpark as the previous unset default).
    cores = os.cpu_count() or 4
    intra_threads = max(1, cores // max(1, workers))
    # 1) Preferred: modern rapidocr (PP-OCRv5, ~13pt more accurate than v4). GPU auto-engages via
    #    onnxruntime-gpu. Guarded by a self-test so a version/API mismatch can't break extraction.
    if os.environ.get("VIEWER_OCR_V5", "1") != "0":
        try:
            from rapidocr import RapidOCR as RapidV5
            # intra_op_num_threads wiring verified against the real `rapidocr` package source (v3.9.2,
            # what `pip install rapidocr` resolves to as of this writing -- downloaded and read in the
            # dev sandbox since the package isn't importable there): rapidocr/main.py's
            # RapidOCR.__init__(config_path=None, params=None) accepts a `params` dict of dotted config
            # keys applied via ParseParams.update_batch() (rapidocr/utils/parse_parameters.py), and
            # rapidocr/inference_engine/onnxruntime/main.py's OrtInferSession._init_sess_opts() reads
            # EngineConfig.onnxruntime.intra_op_num_threads off that same config and sets it on
            # onnxruntime's real SessionOptions -- a verified path, not a guess. Any mismatch here
            # (wrong key, API drift) is caught by the try/except this whole block already runs under,
            # same fail-open-to-v4 behavior as before this change.
            ad = _RapidAdapter(RapidV5(params={
                "EngineConfig.onnxruntime.intra_op_num_threads": intra_threads,
            }), "v5")
            if _selftest(ad):
                log("OCR engine: RapidOCR PP-OCRv5" + (" on GPU" if gpu else " on CPU (intra_op_threads=%d)" % intra_threads))
                return ad
            log("PP-OCRv5 self-test did not extract text; falling back to PP-OCRv4.")
        except Exception as e:
            log("PP-OCRv5 unavailable (%s); using PP-OCRv4." % str(e)[:70])
    # 2) Proven: rapidocr_onnxruntime (PP-OCRv4), with CUDA when requested.
    # NOT given the same intra_op_num_threads treatment as v5 above -- verified (not just unverified/
    # skipped) that this specific wrapper has no hook for it: rapidocr_onnxruntime 1.2.3's session
    # builder (rapidocr_onnxruntime/utils.py, OrtInferSession.__init__) constructs onnxruntime's
    # SessionOptions() with everything hardcoded (log_severity_level, enable_cpu_mem_arena,
    # graph_optimization_level) and never reads ANY thread-count key off its config object at all --
    # confirmed by downloading and reading the actual PyPI wheel in this sandbox (the package itself
    # isn't installed/importable here, so this was checked via its real source, not guessed). Passing
    # e.g. RapidOCR(intra_op_num_threads=N) here would be silently absorbed into the unused Global
    # config dict and do nothing -- worse than leaving it alone -- so it's intentionally not wired.
    from rapidocr_onnxruntime import RapidOCR
    if USE_CUDA:
        try:
            import rapidocr_onnxruntime as _ro
            base = os.path.join(os.path.dirname(_ro.__file__), "config.yaml")
            cfg = open(base, encoding="utf-8").read().replace("use_cuda: false", "use_cuda: true")
            tf = tempfile.NamedTemporaryFile("w", suffix=".yaml", delete=False, encoding="utf-8")
            tf.write(cfg); tf.close()
            log("OCR engine: RapidOCR PP-OCRv4, CUDA requested" + (" (GPU available)" if gpu else " (no GPU -> CPU)"))
            return _RapidAdapter(RapidOCR(config_path=tf.name), "v4")
        except Exception as e:
            log(f"GPU config failed ({str(e)[:70]}); using CPU")
            return _RapidAdapter(RapidOCR(), "v4")
    log("OCR engine: RapidOCR PP-OCRv4 (CPU)")
    return _RapidAdapter(RapidOCR(), "v4")

def log(msg): print(f"[{time.strftime('%H:%M:%S')}] {msg}", flush=True)

def connect(db_path):
    con = sqlite3.connect(db_path, timeout=60)
    con.execute("PRAGMA foreign_keys=ON"); con.execute("PRAGMA synchronous=NORMAL")
    if os.environ.get("VIEWER_RELAXED") == "1":
        con.execute("PRAGMA locking_mode=EXCLUSIVE"); con.execute("PRAGMA journal_mode=TRUNCATE")
    return con

def _sql_statements(script):
    """Split a migration script into individual statements with sqlite3.complete_statement (aware of
    string literals, comments and trigger BEGIN..END bodies), skipping comment-only fragments. Needed
    because executescript() force-COMMITs any open transaction first -- which would break the
    one-transaction-per-migration atomicity guarantee below."""
    stmt = ""
    for line in script.splitlines(keepends=True):
        stmt += line
        if sqlite3.complete_statement(stmt):
            s = stmt.strip()
            if s and any(l.strip() and not l.strip().startswith("--") for l in s.splitlines()):
                yield s
            stmt = ""
    tail = stmt.strip()
    if tail and any(l.strip() and not l.strip().startswith("--") for l in tail.splitlines()):
        yield tail

def migrate(con, migrations_dir, db_path=None):
    """v1.13: each migration's DDL + its schema_version bump commit ATOMICALLY (one BEGIN IMMEDIATE ..
    COMMIT). Previously executescript() committed the DDL first and the version bump after -- a crash
    between the two left columns applied with a stale schema_version, the exact crash-loop class
    fix_schema_version.py exists to patch. Now a crash rolls the whole migration back cleanly.

    Backs up db_path via safeguard.backupdb() before applying any pending migration -- docs/
    ARCHITECTURE.md's standing rule R1 ("no breaking changes without ... a way to roll back")
    applies to schema DDL as much as anything else, and migrations are the least reversible write
    class in this codebase (SQLite's ALTER TABLE ADD COLUMN has no clean per-migration undo; only 1
    of 9 migration files even documents a rollback, and only in a prose comment). The docs used to
    separately claim this backup already happened (a fictional `viewer.db.bak-<version>-<date>`
    file) -- it never did; this makes that claim true, via the same VACUUM INTO backup mechanism
    every other write path that touches viewer.db already uses (ingest_feature.py, run_ocr_auto.bat,
    ENRICH-PUBLOG.bat, ...). Gated on `pending` being non-empty so the common no-op case (nothing to
    migrate, which is every read-only CLI invocation) doesn't pay a multi-GB backup cost. Backup
    failure aborts the whole migrate() call (raises) rather than proceeding without a rollback path
    -- unlike ingest_feature.py's best-effort pre-ingest snapshot, a migration that can't be rolled
    back if it goes wrong is exactly the scenario R1 exists to prevent."""
    has = con.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='schema_meta'").fetchone()
    version = 0
    if has:
        row = con.execute("SELECT schema_version FROM schema_meta WHERE id=1").fetchone()
        version = row[0] if row else 0
    files = sorted(f for f in os.listdir(migrations_dir) if re.match(r"\d{4}_.*\.sql$", f))
    pending = [f for f in files if int(f[:4]) > version]
    if pending and db_path and os.path.exists(db_path):
        import safeguard
        log(f"migrate: {len(pending)} pending migration(s) -- backing up {db_path} before applying...")
        try:
            # Only the CANONICAL index (safeguard.DB_DEFAULT) backs up into the shared, rotated
            # <repo>/backups/db/ vault every other backup path in this codebase already writes to.
            # Any OTHER db_path (a test's tempdir DB, a staging/alternate corpus passed via --db)
            # gets its own sibling backups/db/ folder instead -- review finding: migrating a
            # throwaway/alternate DB must never write into (and keep=2-rotate stale copies out
            # of) the real production backup vault. Confirmed live: running the test suite used
            # to write into and rotate the actual repo's backups/db/ on every run.
            dest_dir = None
            if os.path.abspath(db_path) != os.path.abspath(safeguard.DB_DEFAULT):
                dest_dir = os.path.join(os.path.dirname(os.path.abspath(db_path)), "backups", "db")
            backup_path = safeguard.backupdb(db_path, dest_dir=dest_dir)
            log(f"migrate: backup OK -> {backup_path}")
        except Exception as e:
            log(f"MIGRATION ABORTED: pre-migration backup failed, refusing to apply DDL with no "
                f"rollback path. Fix the backup issue (often: free disk space) and rerun. Error: {e}")
            raise
    applied = 0
    for f in files:
        v = int(f[:4])
        if v > version:
            script = open(os.path.join(migrations_dir, f), encoding="utf-8").read()
            stmts = list(_sql_statements(script))
            # connection-level PRAGMAs (e.g. foreign_keys) are no-ops inside a transaction -> run first
            pragmas = [s for s in stmts if s.lstrip().upper().startswith("PRAGMA")]
            body = [s for s in stmts if s not in pragmas]
            try:
                for s in pragmas:
                    con.execute(s)
                con.commit()                              # close any implicit txn; BEGIN must start clean
                con.execute("BEGIN IMMEDIATE")            # DDL + version bump: ONE atomic unit
                for s in body:
                    con.execute(s)
                con.execute("UPDATE schema_meta SET schema_version=? WHERE id=1", (v,))
                con.commit()
            except Exception as e:
                try: con.rollback()
                except Exception: pass
                log(f"MIGRATION FAILED: {f} -- rolled back ATOMICALLY, schema_version stays {version}; "
                    f"nothing was partially applied. Fix the migration and rerun. Error: {e}")
                raise
            applied += 1; version = v; log(f"applied migration {f}")
    if applied == 0: log(f"schema up to date (v{version})")
    return con

def fingerprint(path, st):
    h = hashlib.md5()
    try:
        with open(path, "rb") as fh: h.update(fh.read(65536))
    except OSError: pass
    return f"{st.st_size}:{int(st.st_mtime)}:{h.hexdigest()[:12]}"

def run_cmd(args, timeout=300):
    return subprocess.run(args, stdout=subprocess.PIPE, stderr=subprocess.DEVNULL, timeout=timeout).stdout.decode("utf-8","ignore")

def pdf_pages_text(path):
    if fitz is not None:
        try:
            doc = fitz.open(path)
            pages = [doc[i].get_text("text") for i in range(doc.page_count)]
            doc.close(); return pages
        except Exception:
            return None
    try: out = run_cmd(["pdftotext","-q",path,"-"], timeout=300)
    except Exception: return None
    return out.split("\f")

def classify_ext(path):
    e = os.path.splitext(path)[1].lower().lstrip(".")
    if e == "pdf": return "pdf"
    if e in ("jpg","jpeg","png","tif","tiff","bmp","gif","svg","svgz"): return "image"
    if e in ("doc","docx","ppt","pptx","xls","xlsx","rtf","txt","htm","html"): return "office"
    return "other"

def upsert_document(con, path, root):
    st = os.stat(path)
    prefix = f"{st.st_size}:{int(st.st_mtime)}"
    row = con.execute("SELECT id, fingerprint FROM documents WHERE path=?", (path,)).fetchone()
    if row and row[1] and row[1].startswith(prefix + ":"):
        return None
    fp = fingerprint(path, st)
    rel = os.path.relpath(path, root)
    vehicle = rel.split(os.sep)[0] if os.sep in rel else rel
    if row and row[1] == fp: return None
    kind = classify_ext(path)
    if row:
        doc_id = row[0]
        con.execute("DELETE FROM pages WHERE document_id=?", (doc_id,))
        con.execute("UPDATE documents SET fingerprint=?, type=?, size_bytes=?, mtime=?, status='discovered', updated_at=datetime('now') WHERE id=?",
                    (fp, kind, st.st_size, st.st_mtime, doc_id))
    else:
        cur = con.execute("INSERT INTO documents(path, rel_path, fingerprint, type, vehicle, size_bytes, mtime, status) VALUES(?,?,?,?,?,?,?, 'discovered')",
                          (path, rel, fp, kind, vehicle, st.st_size, st.st_mtime))
        doc_id = cur.lastrowid
    return doc_id, kind

def index_pdf(con, doc_id, path):
    pages = pdf_pages_text(path)
    if pages is None:
        con.execute("UPDATE documents SET status='failed' WHERE id=?", (doc_id,)); return 0, 0
    indexed = queued = 0; meta_text = ""
    meas_con = _open_meas_db(_db_dir(con))
    try:
        for i, txt in enumerate(pages, start=1):
            body = txt.strip(); cc = len(body)
            if i <= 3: meta_text += " " + body
            if cc < OCR_CHAR_THRESHOLD:
                con.execute("INSERT INTO pages(document_id,page_number,body_text,char_count,source,ocr_status) VALUES(?,?,?,?, 'none','pending')",(doc_id,i,"",cc)); queued += 1
            else:
                con.execute("INSERT INTO pages(document_id,page_number,body_text,char_count,source,ocr_status) VALUES(?,?,?,?, 'text','none')",(doc_id,i,body,cc)); indexed += 1
                if _EXTRACT_TALLY: _EXTRACT_TALLY["dimensions"] += _extract_measures_for_page(meas_con, doc_id, i, body)
    finally:
        if meas_con:
            try: meas_con.close()
            except Exception: pass
    nsn = NSN_RE.search(meta_text); tm = TM_RE.search(meta_text)
    title = next((l.strip() for l in meta_text.splitlines() if len(l.strip()) > 6), "")[:200]
    dtype = "pdf_text" if indexed >= queued else "pdf_scanned"
    con.execute("UPDATE documents SET page_count=?, tm_number=?, nsn=?, title=?, type=?, status=? WHERE id=?",
                (len(pages), tm.group(0) if tm else None, nsn.group(0) if nsn else None, title, dtype, 'indexed' if queued==0 else 'partial', doc_id))
    return indexed, queued

# Discovery Engine, phase 1 -- non-PDF content extraction. Before this, classify_ext() already
# recognized 'image' and 'office' documents (crawl() has always been ABLE to tell a .jpg from a
# .docx), but crawl()'s dispatch did nothing with either: any non-PDF document was just marked
# status='indexed' with zero pages, discovered but never actually read -- confirmed by grep, there
# was no code path at all. This closes that gap for the formats reachable without adding new
# dependencies (images, plain text, HTML); genuine Office-format parsing (.doc/.docx/.ppt/.pptx/
# .xls/.xlsx/.rtf) needs new libraries (python-docx/openpyxl/python-pptx) this pass deliberately
# doesn't add -- those documents are still discovered exactly as before, just still zero pages.
_IMAGE_EXTS = (".jpg", ".jpeg", ".png", ".tif", ".tiff", ".bmp", ".gif")
# .svg/.svgz deliberately excluded: vector MARKUP, not a raster photo -- PyMuPDF can't page-render
# arbitrary SVG the way it can a raster image or PDF, and "parse SVG content" is a different problem
# (structured markup, not pixels to OCR) than every other format handled here.
_TEXT_EXTS = (".txt",)
_HTML_EXTS = (".htm", ".html")
_HTML_TAG_RE = re.compile(r"<script\b.*?</script>|<style\b.*?</style>|<[^>]+>", re.I | re.S)
_HTML_ENTITY_RE = re.compile(r"&(amp|lt|gt|quot|#39|nbsp);")
_HTML_ENTITY_MAP = {"amp": "&", "lt": "<", "gt": ">", "quot": '"', "#39": "'", "nbsp": " "}

def _strip_html(raw):
    """Dependency-free HTML->text: strip <script>/<style> blocks and every tag, then unescape the
    handful of entities plain-text manual exports actually use. Not a real HTML parser (no
    lxml/beautifulsoup4 in this pass's scope -- see the module comment above) -- good enough for a
    structured export/report, not meant to survive adversarial/malformed markup."""
    text = _HTML_TAG_RE.sub(" ", raw or "")
    text = _HTML_ENTITY_RE.sub(lambda m: _HTML_ENTITY_MAP.get(m.group(1), m.group(0)), text)
    return re.sub(r"[ \t]+", " ", text).strip()

def index_other(con, doc_id, path):
    """Non-PDF content extraction for the formats reachable without new dependencies:
      image (.jpg/.png/.tif/...) -- queued for OCR exactly like a scanned PDF page. PyMuPDF opens a
        raw image file as a 1-page document (verified: fitz.open('photo.png').page_count == 1,
        get_pixmap() renders it identically to a real page) -- so _render_png()/ocr_one() need ZERO
        changes to handle it: the ENTIRE existing pipeline (blank-page skip, OCR engine, barcode
        scan, dimensional extraction) runs on a standalone image for free, the moment its `pages`
        row exists with ocr_status='pending'.
      .txt -- read directly as the page's body_text, no OCR needed.
      .htm/.html -- stdlib-only tag-stripped to plain text (_strip_html() above), no OCR needed.
      anything else 'office'-classified (.doc/.docx/.ppt/.pptx/.xls/.xlsx/.rtf) -- discovered, not
        extracted (see module comment above).
    Returns (indexed, queued), same contract index_pdf() already has, so crawl()'s pi/q tallies and
    every caller that unpacks that pair stay correct for these document types too."""
    ext = os.path.splitext(path)[1].lower()
    if ext in _IMAGE_EXTS:
        con.execute("INSERT INTO pages(document_id,page_number,body_text,char_count,source,ocr_status) "
                    "VALUES(?,1,'',0,'none','pending')", (doc_id,))
        con.execute("UPDATE documents SET page_count=1, type='image', status='partial' WHERE id=?", (doc_id,))
        return 0, 1
    if ext in _TEXT_EXTS or ext in _HTML_EXTS:
        try:
            with open(path, "r", encoding="utf-8", errors="ignore") as f:
                raw = f.read(5_000_000 if ext in _HTML_EXTS else 2_000_000)
            body = _strip_html(raw) if ext in _HTML_EXTS else raw.strip()
        except Exception:
            body = ""
        con.execute("INSERT INTO pages(document_id,page_number,body_text,char_count,source,ocr_status) "
                    "VALUES(?,1,?,?, 'text','none')", (doc_id, body, len(body)))
        con.execute("UPDATE documents SET page_count=1, type=?, status='indexed' WHERE id=?",
                    ("html" if ext in _HTML_EXTS else "text", doc_id))
        if body and _EXTRACT_TALLY:
            meas_con = _open_meas_db(_db_dir(con))
            if meas_con:
                try:
                    _EXTRACT_TALLY["dimensions"] += _extract_measures_for_page(meas_con, doc_id, 1, body)
                finally:
                    meas_con.close()
        return (1, 0) if body else (0, 0)
    con.execute("UPDATE documents SET status='indexed' WHERE id=?", (doc_id,))   # unsupported -- unchanged from before this function existed
    return 0, 0

def _track_crawled_doc(con, doc_id):
    """Document metadata + touched-doc-id tracking, shared by crawl()'s PDF and image/text/html
    branches -- WHERE this document's own identity will surface (the home page's 'Browse by
    vehicle' grouping, search result headers, the in-app scan UI's breakdown panel) once its
    indexer has actually determined it, not just 'N files scanned'. Best-effort: a failure here
    must never break ingest."""
    try:
        drow = con.execute(
            "SELECT tm_number, nsn, title, vehicle, page_count, type FROM documents WHERE id=?",
            (doc_id,)).fetchone()
        if drow:
            _EXTRACT_TALLY["documents"].append({
                "id": doc_id, "tm_number": drow[0], "nsn": drow[1], "title": drow[2],
                "vehicle": drow[3], "page_count": drow[4], "type": drow[5]})
    except Exception: pass
    _TOUCHED_DOC_IDS.add(doc_id)   # so the schematics stage (after extract_parts()) considers this document

def crawl(con, root, max_files=0, max_seconds=0):
    rid = con.execute("INSERT INTO runs(kind) VALUES('crawl')").lastrowid
    dbdir = _db_dir(con)
    seen=new=pi=q=fail=0; t0=time.time()
    if not _EXTRACT_TALLY: _tally_reset()   # main() resets per subprocess; direct callers (tests) get a lazy default
    _write_progress(dbdir, stage="crawl", current=None, seen=0, new=0, extracted=dict(_EXTRACT_TALLY))
    for dirpath, dirnames, filenames in os.walk(root, followlinks=True):
        if os.sep + "." in dirpath: continue
        for fn in filenames:
            if fn.lower() in ("thumbs.db",".ds_store"): continue
            path = os.path.join(dirpath, fn); seen += 1
            _write_progress(dbdir, stage="crawl", current=fn, seen=seen, new=new, extracted=dict(_EXTRACT_TALLY))
            try:
                res = upsert_document(con, path, root)
                if res is None:
                    if max_seconds and seen % 200 == 0 and time.time()-t0 > max_seconds:
                        con.commit(); log(f"crawl: time budget reached while scanning (seen={seen}), pausing"); return
                    continue
                doc_id, kind = res; new += 1
                if kind == "pdf":
                    a,b = index_pdf(con, doc_id, path); pi += a; q += b
                    _track_crawled_doc(con, doc_id); _EXTRACT_TALLY["pages_text"] += a
                    _write_progress(dbdir, stage="crawl", current=fn, seen=seen, new=new, extracted=dict(_EXTRACT_TALLY))
                elif kind in ("image", "office"):
                    # Discovery Engine phase 1: images/.txt/.html actually get their content read
                    # now (see index_other()'s own docstring) -- same document-metadata tracking as
                    # the PDF branch above, so they show up correctly in the breakdown panel too.
                    a,b = index_other(con, doc_id, path); pi += a; q += b
                    _track_crawled_doc(con, doc_id); _EXTRACT_TALLY["pages_text"] += a
                    _write_progress(dbdir, stage="crawl", current=fn, seen=seen, new=new, extracted=dict(_EXTRACT_TALLY))
                else:
                    con.execute("UPDATE documents SET status='indexed' WHERE id=?", (doc_id,))
                if new % 10 == 0:
                    con.execute("UPDATE runs SET files_seen=?, new_docs=?, pages_indexed=?, ocr_queued=? WHERE id=?",(seen,new,pi,q,rid)); con.commit()
                    log(f"crawl: seen={seen} new={new} pages_text={pi} ocr_queued={q} ({seen/max(1,time.time()-t0):.0f} files/s)")
                if max_files and new >= max_files:
                    con.execute("UPDATE runs SET finished_at=datetime('now'), files_seen=?, new_docs=?, pages_indexed=?, ocr_queued=?, failed=? WHERE id=?",(seen,new,pi,q,fail,rid)); con.commit()
                    log(f"crawl: batch cap {max_files} reached, pausing (resumable)"); return
                if max_seconds and time.time()-t0 > max_seconds:
                    con.execute("UPDATE runs SET finished_at=datetime('now'), files_seen=?, new_docs=?, pages_indexed=?, ocr_queued=?, failed=? WHERE id=?",(seen,new,pi,q,fail,rid)); con.commit()
                    log(f"crawl: time budget {max_seconds}s reached, pausing (resumable)"); return
            except Exception as e:
                fail += 1; log(f"ERROR {path}: {e}")
    con.execute("UPDATE runs SET finished_at=datetime('now'), files_seen=?, new_docs=?, pages_indexed=?, ocr_queued=?, failed=? WHERE id=?",(seen,new,pi,q,fail,rid)); con.commit()
    log(f"CRAWL DONE seen={seen} new={new} pages_text={pi} ocr_queued={q} failed={fail}")

def _capped_dpi(w_in, h_in, d):
    """Shrink `d` (DPI), never below MIN_OCR_DPI, so a page of physical size (w_in, h_in) inches
    rasterizes at or under OCR_MAX_MEGAPIXELS pixels. Returns `d` unchanged if the page is already
    under the cap at that DPI, or if the dimensions aren't usable (<=0 -- e.g. a malformed
    MediaBox). Shared by both render backends (finding #24 review: the fix originally lived only
    in the PyMuPDF branch, leaving the pdftoppm fallback -- the documented path for machines
    without PyMuPDF, e.g. Windows 7/Vista per docs/SYSTEM-REQUIREMENTS.md -- fully uncapped).
    Review finding: the floor used to be 100 DPI, which for a large enough page (area over
    OCR_MAX_MEGAPIXELS/100**2, e.g. ~2500 sq in at the default 25MP cap) produced a raster WELL
    OVER the cap -- the exact "~60+ megapixel raster" problem this fix exists to prevent, just at
    a higher size threshold. MIN_OCR_DPI is deliberately low (not literally 1) only to avoid a
    truly degenerate near-zero-DPI render on some absurd outlier; the megapixel cap is the hard
    guarantee this function exists to uphold, and always wins over the OCR-quality floor."""
    if w_in > 0 and h_in > 0 and (w_in * d) * (h_in * d) > OCR_MAX_MEGAPIXELS:
        return max(MIN_OCR_DPI, int((OCR_MAX_MEGAPIXELS / (w_in * h_in)) ** 0.5))
    return d


def _pdftoppm_page_size_in(path, page_number):
    """Physical page size in inches via `pdfinfo` (the same Poppler package pdftoppm ships with),
    for the non-PyMuPDF fallback render path. Returns (None, None) on any failure (missing
    pdfinfo, unparseable output, timeout) -- callers must treat that as 'can't determine size,
    render uncapped' (fail-open: a preprocessing quality/safety nicety must never be the reason a
    page fails to render at all)."""
    try:
        out = subprocess.run(["pdfinfo", "-f", str(page_number), "-l", str(page_number), path],
                             stdout=subprocess.PIPE, stderr=subprocess.DEVNULL, timeout=30).stdout.decode("utf-8", "ignore")
        m = re.search(r"Page size:\s*([\d.]+)\s*x\s*([\d.]+)\s*pts", out)
        if m:
            return float(m.group(1)) / 72.0, float(m.group(2)) / 72.0
    except Exception:
        pass
    return None, None


def _render_png(path, page_number, dpi=None):
    d = int(dpi or OCR_DPI)
    if fitz is not None:
        # Bounded acquire, not a bare `with _FITZ_LOCK:` -- PyMuPDF's C state isn't thread-safe, so
        # every render in this process serializes behind this one lock. If a render ever wedges
        # *while holding it*, an unbounded acquire means every OTHER worker thread's next call to
        # _render_png() also blocks forever trying to acquire the same lock -- "one page hangs"
        # silently becomes "the whole process wedges on its very next render." A bounded acquire
        # means a page that can't get the lock in time fails fast (raises, caught by _ocr_task's
        # own timeout/except handling) instead of hanging, so the batch keeps draining even after a
        # wedge -- degraded, but ocrall() still exits normally and the existing crash/restart loop
        # in run_ocr_auto.bat can pick it back up (finding #15).
        if not _FITZ_LOCK.acquire(timeout=OCR_LOCK_TIMEOUT_SECONDS):
            raise TimeoutError("PyMuPDF render lock busy for >%ds -- a prior render may be wedged" % OCR_LOCK_TIMEOUT_SECONDS)
        try:
            doc = fitz.open(path); page = doc[page_number-1]
            # DPI ceiling (finding #24): shrink the effective DPI, never the requested one, for a
            # physically large page so the rendered raster stays under OCR_MAX_MEGAPIXELS.
            w_in, h_in = page.rect.width / 72.0, page.rect.height / 72.0
            d2 = _capped_dpi(w_in, h_in, d)
            if d2 != d:
                log(f"ocr: page {page_number} of {os.path.basename(path)} is {w_in:.1f}x{h_in:.1f}in "
                    f"-- DPI reduced {d}->{d2} to stay under {OCR_MAX_MEGAPIXELS // 1_000_000}MP")
                d = d2
            pix = page.get_pixmap(dpi=d)
            data = pix.tobytes("png"); doc.close()
        finally:
            _FITZ_LOCK.release()
        with tempfile.NamedTemporaryFile(suffix=".png", delete=False) as tf:
            tf.write(data); return tf.name
    w_in, h_in = _pdftoppm_page_size_in(path, page_number)
    if w_in and h_in:
        d2 = _capped_dpi(w_in, h_in, d)
        if d2 != d:
            log(f"ocr: page {page_number} of {os.path.basename(path)} is {w_in:.1f}x{h_in:.1f}in "
                f"-- DPI reduced {d}->{d2} to stay under {OCR_MAX_MEGAPIXELS // 1_000_000}MP (pdftoppm fallback)")
            d = d2
    td = tempfile.mkdtemp(); prefix = os.path.join(td, "pg")
    subprocess.run(["pdftoppm","-r",str(d),"-f",str(page_number),"-l",str(page_number),"-png",path,prefix],
                   stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL, timeout=180)
    pngs = [os.path.join(td,f) for f in os.listdir(td) if f.endswith(".png")]
    if not pngs: raise RuntimeError("pdftoppm produced no image")
    return pngs[0]

def _page_is_blank(path, page_number):
    """Skip-the-junk: True if the page is essentially blank (covers, dividers, empty backs)."""
    if fitz is None or _np is None: return False
    try:
        doc = fitz.open(path)
        pix = doc[page_number-1].get_pixmap(dpi=50, colorspace=fitz.csGRAY)
        arr = _np.frombuffer(pix.samples, dtype=_np.uint8); doc.close()
        if arr.size == 0: return True
        return float((arr < 110).mean()) < 0.004   # < 0.4% dark pixels => blank
    except Exception:
        return False

def _ocr_preprocessed_input(img_path, for_tesseract, _pil_img=None):
    """Best-effort OCR preprocessing (finding #14). Returns what to actually feed to the OCR
    engine: for RapidOCR (for_tesseract=False), a preprocessed numpy array via
    ocrprep.preprocess_light() -- deskew + denoise, NOT binarize, since a hard Otsu threshold is a
    classic-OCR-era optimization that can hurt a deep-learning detector/recognizer by discarding
    anti-aliasing/gradient information a DL model can actually use. For Tesseract (for_tesseract=
    True, this project's fallback engine -- a classic pipeline that DOES benefit from
    binarization), a NEW temp PNG file path via the full ocrprep.preprocess() (deskew + denoise +
    binarize). Falls back to the raw, unmodified image on any failure or when OpenCV/numpy aren't
    available -- this is a quality improvement, never a hard requirement; a preprocessing bug must
    never be the reason a page fails to OCR at all.

    `_pil_img`, if given, is an already-open/RGB-converted PIL Image of img_path's bytes -- reused
    from ocr_one()'s single shared decode (review finding: this and _scan_barcode() were each
    independently opening + RGB-converting the SAME rendered PNG from disk, a second full PIL decode
    of identical bytes on every OCR'd page). Purely an optimization: omit it (default None) and this
    opens img_path itself exactly as before -- callers outside ocr_one() are unaffected."""
    if not (OCR_PREPROCESS and ocrprep is not None and ocrprep.available() and _np is not None):
        return img_path
    try:
        if _pil_img is not None:
            arr = _np.array(_pil_img)
        else:
            from PIL import Image
            with Image.open(img_path) as im:
                arr = _np.array(im.convert("RGB"))
        if for_tesseract:
            import cv2
            proc, _meta = ocrprep.preprocess(arr)
            out_path = img_path + ".prep.png"
            # cv2.imwrite() can fail (bad path, bad array) by returning False, not by raising --
            # review finding: the old unconditional `return out_path` broke the docstring's own
            # "falls back to the raw image on any failure" promise on exactly that failure mode.
            if not cv2.imwrite(out_path, proc):
                return img_path
            return out_path
        proc, _meta = ocrprep.preprocess_light(arr)
        # Review finding: preprocess_light() always collapses to a 2-D grayscale array (ocrprep.
        # _gray() is unconditional inside deskew()/denoise()); RapidOCR is only ever self-tested
        # against a 3-channel array elsewhere in this codebase, so restore 3 channels here rather
        # than ship an unverified grayscale-array shape on this default-on path.
        if getattr(proc, "ndim", 3) == 2:
            import cv2
            proc = cv2.cvtColor(proc, cv2.COLOR_GRAY2RGB)
        return proc
    except Exception:
        return img_path


def _scan_barcode(img_path, _pil_img=None):
    """Best-effort barcode/QR/Data-Matrix read off the SAME rendered page PNG _render_png() already
    produced for OCR (img_path) -- never a second render of the page. Some TMs print NSNs/part
    numbers as barcodes; a machine-decoded value has no character-recognition ambiguity, so it is
    higher-trust provenance than OCR text (barcodes.py catalog §4.9 -- had zero callers before this).

    Opt-in + cheap: returns None immediately if VIEWER_BARCODE_SCAN=0, the module failed to import,
    or barcodes.available() is False (neither pyzbar nor OpenCV installed) -- mirrors barcodes.py's
    own graceful-degradation contract, so this is a true no-op on an environment without the optional
    backends, not a feature that silently requires them. Also swallows any decode failure (a barcode-
    read bug must never break OCR itself -- R1); on success returns the FIRST decoded record as
    {'type','data','nsn'} (nsn omitted/None if the payload has no recognizable NSN), preferring a
    record that DID decode an NSN over one that didn't when a page carries more than one barcode.

    `_pil_img`, if given, is an already-open/RGB-converted PIL Image of img_path's bytes -- reused
    from ocr_one()'s single shared decode (review finding: this and _ocr_preprocessed_input() were
    each independently opening + RGB-converting the SAME rendered PNG from disk, a second full PIL
    decode of identical bytes on every OCR'd page). Purely an optimization: omit it (default None)
    and this opens img_path itself exactly as before -- direct callers (incl. tests) are unaffected."""
    if not BARCODE_SCAN or barcodes is None or not barcodes.available():
        return None
    try:
        if _pil_img is not None:
            recs = barcodes.detect(_pil_img)
        else:
            from PIL import Image
            with Image.open(img_path) as im:
                recs = barcodes.detect(im.convert("RGB"))
    except Exception:
        return None
    if not recs:
        return None
    rec = next((r for r in recs if r.get("nsn")), recs[0])
    return {"type": rec.get("type"), "data": (rec.get("data") or "")[:500], "nsn": rec.get("nsn")}


def ocr_one(path, page_number):
    """Returns (text, confidence, barcode). confidence is RapidOCR's page-level average of its per-line
    detection scores (0.0-1.0), rounded to 4dp -- None on the blank-skip path, the Tesseract fallback (no
    per-line scores exposed the same way), or if RapidOCR returned no scored lines. v1.13.5: this score was
    always being computed (see _RapidAdapter, r[2]) but silently discarded here -- captured now as the first
    real, corpus-wide OCR-quality signal (previously the only signal was 'OCR ran' vs 'OCR did not run').
    barcode is _scan_barcode()'s result -- None, or {'type','data','nsn'} -- read off the SAME render used
    for OCR when the page gets a full OCR pass, or off a dedicated render on the blank-skip path (see
    below); None whenever BARCODE_SCAN/barcodes.py can't run (opt-in + cheap, see _scan_barcode())."""
    dens = _page_density(path, page_number)
    if dens is not None and dens < 0.004:
        # skip-the-junk: no full OCR on blanks (same threshold) -- but a TM divider/parts-label/cover
        # page is routinely near-empty EXCEPT for a small barcode/QR stamp, i.e. exactly the page most
        # likely to fall under this OCR-tuned density threshold and most likely to carry a machine-
        # readable NSN. Render + scan for a barcode before giving up on the page entirely, gated the
        # same way _scan_barcode() itself is gated so a render is never paid for here when barcode
        # scanning is off or unavailable.
        barcode = None
        if BARCODE_SCAN and barcodes is not None and barcodes.available():
            bimg = None
            try:
                bimg = _render_png(path, page_number, OCR_DPI)
                barcode = _scan_barcode(bimg)
            except Exception:
                barcode = None
            finally:
                if bimg and os.path.exists(bimg):
                    try: os.unlink(bimg)
                    except OSError: pass
        return "", None, barcode
    dpi = OCR_DPI
    if ADAPTIVE_DPI and dens is not None and dens < 0.02:   # opt-in: sparse pages render lower (never below 160)
        dpi = max(160, OCR_DPI - 50)
    img = None
    tess_img = None
    try:
        img = _render_png(path, page_number, dpi)
        h = None                                          # identical-page dedup: reuse text for repeated boilerplate
        try:
            with open(img, "rb") as f: h = hashlib.md5(f.read()).hexdigest()
        except Exception: h = None
        if h is not None:
            with _DEDUP_LOCK: cached = _DEDUP.get(h)
            if cached is not None:
                _DEDUP_STATS["hits"] += 1; return cached
        # Dedup hash above is deliberately over the RAW render, before preprocessing -- identical
        # source pages still produce an identical raw render, so the dedup key is unaffected by
        # whether preprocessing itself is on/off or changes over time. Barcode scan runs against
        # that same raw render too (before it goes into the cache), so an identical repeated page
        # (boilerplate covers, dividers) reuses the barcode read exactly like it reuses OCR text.
        #
        # Review finding: _scan_barcode() and _ocr_preprocessed_input() each independently opened +
        # RGB-converted this SAME PNG from disk -- a second full PIL decode of identical bytes on
        # every OCR'd page (never a second RENDER, which the docstrings above correctly guard against,
        # but still real, avoidable per-page decode work across a whole corpus). Decode once here and
        # hand the shared image to both; best-effort only -- if PIL isn't importable or this decode
        # fails, shared_rgb stays None and each callee falls back to its own independent open(),
        # exactly as before.
        shared_rgb = None
        try:
            from PIL import Image
            with Image.open(img) as im:
                shared_rgb = im.convert("RGB")
        except Exception:
            shared_rgb = None
        barcode = _scan_barcode(img, _pil_img=shared_rgb)
        try:
            if _have_rapid():
                ocr_input = _ocr_preprocessed_input(img, for_tesseract=False, _pil_img=shared_rgb)
                res, _ = _get_rapid()(ocr_input)
                text = "\n".join(r[1] for r in res).strip() if res else ""
                scores = [r[2] for r in res if len(r) > 2 and isinstance(r[2], (int, float))] if res else []
                conf = round(sum(scores) / len(scores), 4) if scores else None
            else:
                tess_img = _ocr_preprocessed_input(img, for_tesseract=True, _pil_img=shared_rgb)
                out = subprocess.run(["tesseract", tess_img, "-", "-l", "eng", "--psm", "1"],
                                     stdout=subprocess.PIPE, stderr=subprocess.DEVNULL, timeout=180)
                text = out.stdout.decode("utf-8","ignore").strip()
                conf = None   # tesseract fallback: no per-line confidence captured (yet)
        except Exception as e:
            # The OCR text engine (RapidOCR or the tesseract binary) failing is a REAL page-level
            # failure and must still surface as one -- ocr()'s handle() marks the page 'failed' and
            # queues a retry job on this, same as always (e.g. tesseract genuinely missing from PATH
            # should be loud, not silently swallowed into an empty-but-'done' page). But `barcode` was
            # already independently decoded above -- a self-contained OpenCV/pyzbar read that has
            # nothing to do with the text engine -- so it must NOT be discarded as collateral damage
            # of an unrelated engine failure. Re-raising a bare exception here would lose `barcode`
            # entirely: ocr_one()'s 3-tuple return never happens, so the caller has nothing to recover
            # it from. Attach it to the exception instead so _ocr_task() can still pull it off the
            # failure path and hand it to handle(), which persists it alongside the 'failed' status.
            e.barcode = barcode
            raise
        result = (text, conf, barcode)
        if h is not None:
            with _DEDUP_LOCK:
                if len(_DEDUP) < 200000: _DEDUP[h] = result
        return result
    finally:
        if img and os.path.exists(img):
            try: os.unlink(img)
            except OSError: pass
        if tess_img and tess_img != img and os.path.exists(tess_img):
            try: os.unlink(tess_img)
            except OSError: pass

def _ocr_task(args):
    """Runs ocr_one() with a wall-clock deadline (OCR_PAGE_TIMEOUT_SECONDS) via a helper thread +
    .join(timeout) -- see the module-level comment by OCR_PAGE_TIMEOUT_SECONDS for why this, not
    signal.alarm or a bare subprocess. Called both from ocr()'s single-threaded loop (workers<=1)
    and from inside its ThreadPoolExecutor (workers>1) -- wrapping it here (rather than in ocr()
    itself) means the timeout applies uniformly in both cases with one implementation. Times out ->
    same (pid, None, None, None, err) shape as any other failure, so handle() needs no changes: the
    page is marked 'failed' and the batch moves on to the next one instead of hanging on it forever.
    args[:3] (not a bare 3-way unpack): ocr()'s `rows` query gained a 4th column (p.document_id,
    for the measures.db/schematics-stage document tracking added alongside it) that this function
    itself has no use for -- only handle()/_labels do."""
    pid, pno, path = args[:3]
    box = {}
    def run():
        try:
            box["text"], box["conf"], box["barcode"] = ocr_one(path, pno)
        except Exception as e:
            box["err"] = str(e)[:300]
            box["barcode"] = getattr(e, "barcode", None)   # see ocr_one()'s except clause: a text-engine
            # failure (e.g. tesseract missing) doesn't mean the independently-decoded barcode is gone too.
    t = threading.Thread(target=run, daemon=True)
    t.start()
    t.join(OCR_PAGE_TIMEOUT_SECONDS)
    if t.is_alive():
        return pid, None, None, None, "timeout after %ds" % OCR_PAGE_TIMEOUT_SECONDS
    if "err" in box:
        return pid, None, None, box.get("barcode"), box["err"]
    return pid, box.get("text"), box.get("conf"), box.get("barcode"), None

def _db_dir(con):
    try:
        for _seq, name, fpath in con.execute("PRAGMA database_list"):
            if name == "main" and fpath: return os.path.dirname(os.path.abspath(fpath))
    except Exception: pass
    return None

def _heartbeat(d, done, fail, remaining):
    """Write a progress heartbeat next to the index so a watchdog can tell a working pass from a hung
    one. Best-effort: a write failure never affects OCR."""
    if not d: return
    try:
        with open(os.path.join(d, "ocr_heartbeat.txt"), "w") as f:
            f.write("%s done=%d fail=%d remaining=%s\n" % (time.strftime("%Y-%m-%d %H:%M:%S"), done, fail, remaining))
    except Exception: pass

def _write_progress(d, **fields):
    """Best-effort JSON progress sidecar (ingest_progress.json, next to the index) for the in-app
    'Add documents' UI (features/ingest_feature.py's ingest_status() reads it) -- deliberately
    separate from _heartbeat()'s plain-text file above, whose exact format ocr_watchdog.py/
    ocr_supervisor.py depend on for stale-hang detection and must never change. This is purely
    additive, UI-facing progress state: which stage the run/ocrall pipeline is in (crawl/ocr/parts/
    done), what item it's on right now (a filename, or a {'doc','page'} pair), and done/total counts
    for whichever stage is active. Atomic write (temp file + os.replace) so a concurrent reader (the
    HTTP status route, polled every ~2s from the browser) never sees a half-written file -- same
    pattern safeguard.py's atomic builds use for the same reason."""
    if not d: return
    try:
        os.makedirs(d, exist_ok=True)
        path = os.path.join(d, "ingest_progress.json")
        tmp = path + ".tmp.%d" % os.getpid()
        with open(tmp, "w") as f:
            json.dump({"updated": time.time(), **fields}, f)
        os.replace(tmp, path)
    except Exception: pass


# ---- extraction tally: WHERE each bit of parsed data actually goes, live -----------------------
# A stage/done/total progress bar answers "how far along is this" but not "what did it actually
# find, and which part of the app will it show up in" -- real, distinguishable data types (a
# document's own metadata, searchable page text, structured NSN/part records, machine-read
# barcodes, OCR confidence) fan out to different consumers (search, /api/part, the home page's
# vehicle browser, quality flags elsewhere) and the UI has no way to show that breakdown without
# this. One process-lifetime accumulator, reset once per subprocess invocation (main(), before
# dispatch) -- crawl()/ocr()/extract_parts() each add to it and include the WHOLE running total in
# every _write_progress() call, so the browser always sees the full picture regardless of which
# stage is currently active, not just that one function's local counters.
_EXTRACT_TALLY = {}
# Which document ids this process has actually touched (crawl discovered, or ocr() OCR'd a page of)
# this run -- deliberately NOT part of _EXTRACT_TALLY itself (a plain set isn't JSON-serializable,
# and _write_progress() dumps a shallow copy of the tally on every call; this is internal bookkeeping
# for the schematics stage below, not UI-facing data). Reset alongside the tally.
_TOUCHED_DOC_IDS = set()
def _tally_reset():
    global _EXTRACT_TALLY, _TOUCHED_DOC_IDS
    _EXTRACT_TALLY = {"documents": [], "pages_text": 0, "pages_ocr_done": 0, "pages_ocr_fail": 0,
                       "barcodes_decoded": 0, "parts_page": 0, "parts_barcode": 0, "nsn_samples": [],
                       "ocr_conf_sum": 0.0, "ocr_conf_n": 0, "dimensions": 0, "schematics": 0, "tables": 0}
    _TOUCHED_DOC_IDS = set()


# ---- dimensional data, live: measures.py + specparse.py + leadingspecs.py, called inline per page
# instead of the separate BUILD-MEASURES.bat batch pass. Deliberately writes ONLY the per-(doc,page)
# layer (measures.db's `meas` table -- the exact schema build_measures.py already defines) and never
# touches masterfile.db's cross-document aggregation (masterfile.build() groups by vehicle label and
# discards the page citation entirely -- explicitly out of scope here; that stays a separate, human-
# triggered step via BUILD-MASTERFILE.bat, unchanged). Every row this writes is traceable to the
# exact document + page it came from, matching the "no cross-referencing, only what's correlative to
# the part's own document" requirement this exists to satisfy.
_MEAS_SCHEMA = """
CREATE TABLE IF NOT EXISTS meas(
  id INTEGER PRIMARY KEY, doc INTEGER, page INTEGER, type TEXT, unit TEXT,
  value TEXT, value2 TEXT, tolerance TEXT, raw TEXT, context TEXT);
CREATE INDEX IF NOT EXISTS ix_meas_doc  ON meas(doc);
CREATE INDEX IF NOT EXISTS ix_meas_type ON meas(type);
"""

def _open_meas_db(dbdir):
    """Best-effort measures.db connection, same schema/location build_measures.py already uses
    (index/measures.db, next to viewer.db) -- so the existing /api/measures, /master, /mastercov
    routes and BUILD-MASTERFILE.bat all pick up live-extracted rows with zero changes of their own.
    Returns None (never raises) on any failure, OR instantly when MEASURES_SCAN is off -- dimensional
    extraction is enrichment, not core ingest, and must never be the reason a scan job fails or slows
    down for an operator who's opted out."""
    if not dbdir or not MEASURES_SCAN: return None
    try:
        os.makedirs(dbdir, exist_ok=True)
        mcon = sqlite3.connect(os.path.join(dbdir, "measures.db"))
        mcon.executescript(_MEAS_SCHEMA)
        return mcon
    except Exception:
        return None

def _extract_measures_for_page(meas_con, doc_id, page_number, body_text):
    """Runs measures.py (generic dimensions) + leadingspecs.py (labelled 'Length: 180 in' specs) +
    specparse.py (thread/fit-class/diameter-tolerance/MIL-STD/fluid) over ONE page's text and writes
    every match into meas_con, tied to (doc_id, page_number). Mirrors build_measures.py's per-page
    inner loop (lines 46-60) exactly, just invoked live per page instead of in a separate whole-
    corpus pass. Returns the count written (0 on any failure or empty input) -- fed into
    _EXTRACT_TALLY['dimensions'] by the caller, same "where did my data go" breakdown the scan UI
    already shows for pages/parts/barcodes."""
    if not meas_con or not body_text:
        return 0
    try:
        import measures as _measures
        rows = _measures.extract(body_text, page=page_number, cap=120)
        try:
            import leadingspecs as _leadingspecs
            rows = rows + _leadingspecs.as_measurements(body_text, page=page_number)
        except Exception:
            pass
        try:
            import specparse as _specparse
            # specparse's shape ({kind,value,context,[tolerance],[page]}) differs slightly from
            # measures.py's ({type,unit,value,...}) -- kind->type, no natural unit (a thread callout
            # or MIL-STD number isn't a magnitude+unit pair the way a length is), raw=value (the
            # match text itself, same convention measures.py's own 'raw' field follows).
            for r in _specparse.extract(body_text, page=page_number, cap=80):
                rows.append({"type": r["kind"], "unit": None, "value": r["value"], "value2": None,
                             "tolerance": r.get("tolerance"), "raw": r["value"], "context": r.get("context")})
        except Exception:
            pass
        if not rows:
            return 0
        meas_con.executemany(
            "INSERT INTO meas(doc,page,type,unit,value,value2,tolerance,raw,context) VALUES(?,?,?,?,?,?,?,?,?)",
            [(doc_id, page_number, m["type"], m.get("unit"), m["value"], m.get("value2"),
              m.get("tolerance"), m.get("raw"), m.get("context")) for m in rows])
        meas_con.commit()
        return len(rows)
    except Exception:
        return 0

def ocr(con, limit, workers=1):
    rid = con.execute("INSERT INTO runs(kind) VALUES('ocr')").lastrowid
    done=fail=0
    dbdir = _db_dir(con)
    # disk guard (RPS-safe): if the index drive is low on space, pause this pass cleanly instead of
    # filling the disk. Fail-open if free space can't be read. The auto-runner will retry.
    if dbdir:
        try:
            from preflight import disk_ok
            ok, free, thr = disk_ok(dbdir)
            if not ok:
                log("ocr: PAUSED -- low disk (%s MB free, need >= %s). Free space and it resumes." % (free, thr))
                con.execute("UPDATE runs SET finished_at=datetime('now') WHERE id=?", (rid,)); con.commit()
                return con.execute("SELECT COUNT(*) FROM pages WHERE ocr_status='pending'").fetchone()[0]
        except Exception:
            pass
    rows = con.execute("SELECT p.id, p.page_number, d.path, p.document_id FROM pages p JOIN documents d ON d.id=p.document_id WHERE p.ocr_status='pending' ORDER BY p.ocr_priority, p.id LIMIT ?", (limit,)).fetchall()
    log(f"ocr: {len(rows)} pages to process this batch (threads={workers}, engine={'RapidOCR' if _have_rapid() else 'tesseract'})")
    if not rows:
        con.execute("UPDATE runs SET finished_at=datetime('now') WHERE id=?", (rid,)); con.commit()
        _heartbeat(dbdir, 0, 0, 0); return 0
    con.executemany("UPDATE pages SET ocr_status='running' WHERE id=?", [(r[0],) for r in rows]); con.commit()
    if _have_rapid(): _get_rapid(workers)      # build the shared engine once, up front
    total = len(rows)
    # pid -> (page_number, doc basename, document_id), so handle() (which only gets a bare pid back
    # from _ocr_task()) can still report a human-readable "currently processing" line for the in-app
    # scan UI (ingest_status() surfaces this via _write_progress()) AND tie a page's OCR'd text back
    # to its document for measures.db (dimensional data is meaningless without knowing which
    # document/page it came from).
    _labels = {r[0]: (r[1], os.path.basename(r[2] or ""), r[3]) for r in rows}
    if not _EXTRACT_TALLY: _tally_reset()   # same lazy-default as crawl(): ocrall doesn't call crawl() first
    meas_con = _open_meas_db(dbdir)
    _write_progress(dbdir, stage="ocr", current=None, done=0, fail=0, total=total, extracted=dict(_EXTRACT_TALLY))
    def handle(pid, text, conf, barcode, err):
        nonlocal done, fail
        if err is None:
            bc = barcode or {}
            con.execute("UPDATE pages SET body_text=?, char_count=?, source='ocr', ocr_status='done', ocr_confidence=?, "
                        "barcode_type=?, barcode_data=?, barcode_nsn=? WHERE id=?",
                        (text, len(text), conf, bc.get("type"), bc.get("data"), bc.get("nsn"), pid)); done += 1
            _EXTRACT_TALLY["pages_ocr_done"] += 1
            if conf is not None:
                _EXTRACT_TALLY["ocr_conf_sum"] += conf; _EXTRACT_TALLY["ocr_conf_n"] += 1
            _lbl_pno, _lbl_dname, _lbl_doc_id = _labels.get(pid, (None, "", None))
            if _lbl_doc_id is not None:
                _EXTRACT_TALLY["dimensions"] += _extract_measures_for_page(meas_con, _lbl_doc_id, _lbl_pno, text)
                _TOUCHED_DOC_IDS.add(_lbl_doc_id)   # so the schematics stage (after extract_parts()) scans this document
        else:
            # The page's OCR text pass genuinely failed (still marked 'failed' + queued for retry,
            # same as always) -- but a barcode reached this far only if it decoded successfully BEFORE
            # the text engine blew up (see ocr_one()'s except clause), so it's real and worth keeping;
            # an OCR-engine outage (e.g. tesseract missing) shouldn't also erase an unrelated,
            # already-good machine-read NSN.
            bc = barcode or {}
            con.execute("UPDATE pages SET ocr_status='failed', barcode_type=?, barcode_data=?, barcode_nsn=? WHERE id=?",
                        (bc.get("type"), bc.get("data"), bc.get("nsn"), pid))
            con.execute("INSERT INTO jobs(page_id,stage,state,attempts,last_error) VALUES(?,?, 'failed', 1, ?)", (pid,"ocr",err)); fail += 1
            _EXTRACT_TALLY["pages_ocr_fail"] += 1
        if bc.get("type"):
            _EXTRACT_TALLY["barcodes_decoded"] += 1   # promoted to a 'barcode'-confidence part later IFF it has an nsn (extract_parts() tallies that half)
        pno, dname, _ = _labels.get(pid, (None, "", None))
        _write_progress(dbdir, stage="ocr", current={"doc": dname, "page": pno}, done=done, fail=fail, total=total,
                        extracted=dict(_EXTRACT_TALLY))
        if (done+fail) % 5 == 0: con.commit(); _heartbeat(dbdir, done, fail, None); log(f"ocr: done={done} failed={fail} (last page {len(text) if text else 0} chars)")
    try:
        if workers <= 1:
            for r in rows: handle(*_ocr_task(r))
        else:
            with ThreadPoolExecutor(max_workers=workers) as ex:
                for fut in as_completed([ex.submit(_ocr_task, r) for r in rows]): handle(*fut.result())
    finally:
        if meas_con:
            try: meas_con.close()
            except Exception: pass
    con.execute("UPDATE documents SET status='indexed' WHERE status='partial' AND id NOT IN (SELECT document_id FROM pages WHERE ocr_status IN ('pending','running','failed'))")
    con.execute("UPDATE runs SET finished_at=datetime('now'), ocr_done=?, failed=? WHERE id=?", (done,fail,rid)); con.commit()
    remaining = con.execute("SELECT COUNT(*) FROM pages WHERE ocr_status='pending'").fetchone()[0]
    _heartbeat(dbdir, done, fail, remaining)
    log(f"OCR BATCH DONE done={done} failed={fail} dedup_reused={_DEDUP_STATS['hits']} remaining_pending={remaining}")
    return remaining

def cleanup(con):
    # Remove leftover sandbox/Unix-path documents (start with '/') whose files don't exist
    # on this machine, and requeue any pages stuck in 'failed'/'running'. Clear non-cascading
    # references (request_items, figures, parts -- none of these cascade-delete, see prune()'s
    # docstring below for why) first so the document delete doesn't trip a FK.
    con.execute("UPDATE request_items SET source_document_id=NULL WHERE source_document_id IN (SELECT id FROM documents WHERE path LIKE '/%')")
    con.execute("DELETE FROM figures WHERE document_id IN (SELECT id FROM documents WHERE path LIKE '/%')")
    # Review finding: this used to omit `parts`, unlike prune()'s otherwise-identical delete
    # sequence below -- orphaning parts rows for every orphan document this removed.
    con.execute("DELETE FROM parts WHERE document_id IN (SELECT id FROM documents WHERE path LIKE '/%')")
    n = con.execute("DELETE FROM documents WHERE path LIKE '/%'").rowcount
    con.commit()
    r = con.execute("UPDATE pages SET ocr_status='pending' WHERE ocr_status IN ('failed','running')").rowcount
    con.execute("DELETE FROM jobs WHERE stage='ocr'")
    con.commit()
    print(f"cleanup: removed {n} orphan documents; requeued {r} pages to 'pending'")

def _prune_sidecars(index_dir, doc_ids, log=log):
    """Best-effort cleanup of the optional per-document sidecars after documents are pruned from the
    main index. Every sidecar here is built by a SEPARATE host-run script (BUILD-MEASURES.bat etc.)
    and may simply not exist yet -- that is not an error, just nothing to prune. A sidecar that DOES
    exist but is locked/corrupt must never abort the (already-committed) main-index prune, so every
    step is wrapped and merely logged on failure (R1: this is cleanup, not a required step)."""
    if not index_dir or not doc_ids:
        return
    ids = list(doc_ids)
    # measures.db: doc-keyed rows, explicit DELETE WHERE doc=? (finding #11's own plan)
    meas_db = os.path.join(index_dir, "measures.db")
    if os.path.exists(meas_db):
        try:
            mc = sqlite3.connect(meas_db, timeout=30)
            try:
                qs = ",".join("?" * len(ids))
                mc.execute("DELETE FROM meas WHERE doc IN (%s)" % qs, ids)
                mc.execute("DELETE FROM meas_done WHERE doc IN (%s)" % qs, ids)
                mc.commit()
            finally:
                mc.close()
        except Exception as e:
            log(f"prune: measures.db cleanup skipped ({e})")
    # tables.db: same doc-keyed shape, same treatment (schema is build_tables.py's own: tables
    # `tbl`/`tbl_done`, both doc-keyed -- confirmed against build_tables.py's SCHEMA, NOT
    # guessed; a review finding caught an earlier version of this using wrong table names,
    # which made the DELETE a silent permanent no-op).
    tab_db = os.path.join(index_dir, "tables.db")
    if os.path.exists(tab_db):
        try:
            tc = sqlite3.connect(tab_db, timeout=30)
            try:
                qs = ",".join("?" * len(ids))
                names = {r[0] for r in tc.execute("SELECT name FROM sqlite_master WHERE type='table'")}
                for tbl, col in (("tbl", "doc"), ("tbl_done", "doc")):
                    if tbl in names:
                        tc.execute("DELETE FROM %s WHERE %s IN (%s)" % (tbl, col, qs), ids)
                tc.commit()
            finally:
                tc.close()
        except Exception as e:
            log(f"prune: tables.db cleanup skipped ({e})")
    # Cache directories keyed by doc_id, one prefix pattern per directory's OWN naming convention
    # (confirmed against each cache's cache_path()/cache_key(), not assumed uniform -- a review
    # finding caught an earlier version of this missing pagecache/ entirely, since rps.py's
    # pagecache uses a HYPHEN ("<doc>-<page>-d<dpi>.png"), not the underscore the other three
    # caches use, plus figcache's callout_crop() output ("callout_<doc>_<page>_<item>_<dpi>.png")
    # doesn't start with the doc_id at all).
    _CACHE_DIR_PREFIXES = {
        "figcache":   ("%d_", "callout_%d_"),
        "schemcache": ("%d_",),
        "veccache":   ("%d_",),
        "pagecache":  ("%d-",),
    }
    for sub, templates in _CACHE_DIR_PREFIXES.items():
        d = os.path.join(index_dir, sub)
        if not os.path.isdir(d):
            continue
        prefixes = tuple(t % i for i in ids for t in templates)
        try:
            for fn in os.listdir(d):
                if fn.startswith(prefixes):
                    try: os.remove(os.path.join(d, fn))
                    except OSError: pass
        except Exception as e:
            log(f"prune: {sub}/ cleanup skipped ({e})")


def prune(con, root="", confirm=False, missing_threshold=0.5, index_dir=None):
    """Reconcile documents whose SOURCE FILE has been deleted or renamed/moved off disk since the
    last crawl (finding #11). crawl()/upsert_document() only ever ADD or UPDATE documents matched by
    exact path -- nothing ever notices a source file that is simply gone, so its row (and every page,
    part, and figure cited from it) sits in the index forever, dead weight that can also surface as a
    broken 'open the PDF' link in the UI. Dry-run by default (same --yes-to-confirm shape as
    rollback()); a plain re-run after --yes shows nothing left to prune.

    Safety valves (a prune is a DELETE -- must never mistake 'can't see the corpus right now' for
    'the corpus shrank'):
      * corpus root unreachable: if `root` is given and isn't a directory (e.g. an unmounted external
        drive or an unmapped network share), abort without looking at a single document -- otherwise
        EVERY document would appear missing and a naive prune would wipe the whole index.
      * missing-fraction threshold: even with a reachable root, if the fraction of indexed documents
        whose file is gone exceeds `missing_threshold` (default 50%), abort and ask for confirmation
        via a higher --missing-threshold -- a mass-disappearance is far more likely a moved corpus
        root or a bad --root than a real mass-deletion.

    A 'renamed/moved' file is detected without re-reading anything: if a missing document's stored
    fingerprint (size:mtime:headhash) matches another, PRESENT document's fingerprint, crawl() has
    already re-discovered that same file under its new path -- the stale old row is a pure duplicate
    and is safe to drop outright (its content is already indexed under the new row).

    Removes the stale row's non-cascading references first (figures, parts: no ON DELETE CASCADE;
    request_items: SET NULL, keeping the session's own history instead of erasing it) so the
    documents DELETE itself doesn't trip a foreign-key error -- pages/jobs DO cascade (schema-level
    ON DELETE CASCADE) and need no manual handling. Then best-effort-prunes the optional per-document
    sidecars (measures.db/tables.db/figcache/schemcache/veccache/pagecache) for the same doc_ids.

    Returns a summary dict; never raises for an ordinary abort (aborted reasons are returned, not
    thrown) so callers/tests can assert on the outcome instead of parsing log output."""
    if root and not os.path.isdir(root):
        log(f"prune: ABORT -- corpus root not reachable ({root!r}); refusing to treat every indexed "
            f"document as deleted. Fix --root (or unmount/mount the drive) and retry.")
        return {"ok": False, "aborted": "root_unreachable", "root": root}

    rows = con.execute("SELECT id, path, fingerprint FROM documents").fetchall()   # tuple rows (this
    total = len(rows)                                                              # file's convention --
    if total == 0:                                                                 # connect() sets no
        log("prune: no documents in the index -- nothing to do")                  # row_factory)
        return {"ok": True, "total": 0, "missing": 0, "deleted": 0, "renamed": 0, "removed_ids": []}

    missing = [r for r in rows if not os.path.isfile(r[1])]
    if not missing:
        log(f"prune: all {total} indexed documents' files exist -- nothing to do")
        return {"ok": True, "total": total, "missing": 0, "deleted": 0, "renamed": 0, "removed_ids": []}

    frac = len(missing) / total
    if frac > missing_threshold:
        log(f"prune: ABORT -- {len(missing)}/{total} ({frac:.0%}) of indexed documents appear missing, "
            f"over the {missing_threshold:.0%} safety threshold. This looks like a moved corpus root or "
            f"an unmounted drive, not a real mass-deletion. Re-run with --root pointing at the corpus, "
            f"or a higher --missing-threshold if this is genuinely intended.")
        return {"ok": False, "aborted": "missing_fraction", "total": total, "missing": len(missing),
                "fraction": round(frac, 4)}

    present_fp = {r[2]: r[0] for r in rows if r[2] and os.path.isfile(r[1])}
    renamed, deleted = [], []
    for r in missing:
        dup = present_fp.get(r[2]) if r[2] else None
        (renamed if dup else deleted).append((r[0], r[1], dup))

    if not confirm:
        log(f"prune: DRY RUN -- would remove {len(missing)}/{total} document(s) "
            f"({len(renamed)} moved/renamed, {len(deleted)} deleted). Re-run with --yes to actually prune.")
        for doc_id, path, dup in (renamed + deleted)[:25]:
            tag = f"moved -> now doc {dup}" if dup else "deleted, no replacement found"
            log(f"  [{tag}] doc {doc_id}: {path}")
        return {"ok": True, "total": total, "missing": len(missing), "deleted": len(deleted),
                "renamed": len(renamed), "removed_ids": [d for d, _, _ in renamed + deleted]}

    ids = [d for d, _, _ in renamed + deleted]
    qs = ",".join("?" * len(ids))
    con.execute(f"UPDATE request_items SET source_document_id=NULL WHERE source_document_id IN ({qs})", ids)
    con.execute(f"DELETE FROM figures WHERE document_id IN ({qs})", ids)
    con.execute(f"DELETE FROM parts WHERE document_id IN ({qs})", ids)
    con.execute(f"DELETE FROM documents WHERE id IN ({qs})", ids)   # pages/jobs cascade (schema ON DELETE CASCADE)
    con.commit()
    log(f"PRUNE DONE: removed {len(ids)} document(s) ({len(renamed)} moved/renamed, {len(deleted)} deleted) "
        f"+ their figures/parts; {len(renamed)} request_items reference(s) cleared to NULL.")
    _prune_sidecars(index_dir, ids, log=log)
    return {"ok": True, "total": total, "missing": len(missing), "deleted": len(deleted),
            "renamed": len(renamed), "removed_ids": ids}


def prioritize(con):
    """Set OCR order: parts catalogs first, then maintenance/troubleshooting, then operator, then rest."""
    try:
        con.execute("""UPDATE pages SET ocr_priority = (SELECT CASE
              WHEN upper(d.path) LIKE '%24P%' OR upper(d.path) LIKE '%PARTS%' OR upper(d.path) LIKE '%RPSTL%' OR upper(coalesce(d.tm_number,'')) LIKE '%24P%' THEN 0
              WHEN upper(d.path) LIKE '%TROUBLESHOOT%' THEN 1
              WHEN upper(d.path) LIKE '%MAINT%' OR upper(coalesce(d.tm_number,'')) LIKE '%-20%' OR upper(coalesce(d.tm_number,'')) LIKE '%-24%' THEN 2
              WHEN upper(d.path) LIKE '%OPERATOR%' OR upper(coalesce(d.tm_number,'')) LIKE '%-10%' THEN 3
              ELSE 5 END FROM documents d WHERE d.id = pages.document_id)
            WHERE ocr_status='pending'""")
        con.commit()
        log("prioritized OCR queue (parts > maintenance/troubleshooting > operator > rest)")
    except sqlite3.OperationalError as e:
        log(f"prioritize skipped ({e}) -- run migrate first")

def prefilter(con, limit, workers=1):
    """Mark blank pending pages as 'skipped' so they leave the OCR queue (shrinks the backlog)."""
    rows = con.execute("SELECT p.id, p.page_number, d.path FROM pages p JOIN documents d ON d.id=p.document_id WHERE p.ocr_status='pending' ORDER BY p.ocr_priority, p.id LIMIT ?", (limit,)).fetchall()
    log(f"prefilter: scanning {len(rows)} pending pages for blanks")
    skipped=0
    for pid, pno, path in rows:
        if _page_is_blank(path, pno):
            con.execute("UPDATE pages SET ocr_status='skipped' WHERE id=?", (pid,)); skipped += 1
            if skipped % 50 == 0: con.commit(); log(f"prefilter: skipped {skipped} blanks")
    con.commit()
    remaining = con.execute("SELECT COUNT(*) FROM pages WHERE ocr_status='pending'").fetchone()[0]
    log(f"PREFILTER DONE skipped={skipped} remaining_pending={remaining}")
    return remaining

def status(con):
    def one(q,*a): return con.execute(q,a).fetchone()[0]
    print("documents       :", one("SELECT COUNT(*) FROM documents"))
    print("  pdf_text      :", one("SELECT COUNT(*) FROM documents WHERE type='pdf_text'"))
    print("  pdf_scanned   :", one("SELECT COUNT(*) FROM documents WHERE type='pdf_scanned'"))
    print("  other types   :", one("SELECT COUNT(*) FROM documents WHERE type NOT LIKE 'pdf%'"))
    print("pages total     :", one("SELECT COUNT(*) FROM pages"))
    print("  text-indexed  :", one("SELECT COUNT(*) FROM pages WHERE source='text'"))
    print("  ocr-indexed   :", one("SELECT COUNT(*) FROM pages WHERE source='ocr'"))
    print("  ocr pending   :", one("SELECT COUNT(*) FROM pages WHERE ocr_status='pending'"))
    print("  ocr failed    :", one("SELECT COUNT(*) FROM pages WHERE ocr_status='failed'"))

def search(con, query, k=8):
    rows = con.execute("SELECT d.vehicle, d.tm_number, p.page_number, snippet(pages_fts,0,'[',']','...',8) FROM pages_fts JOIN pages p ON p.id=pages_fts.rowid JOIN documents d ON d.id=p.document_id WHERE pages_fts MATCH ? ORDER BY rank LIMIT ?", (query,k)).fetchall()
    for v,tm,pg,sn in rows: print(f"  [{v}] {tm or ''} p.{pg}: {sn.replace(chr(10),' ')}")
    if not rows: print("  (no matches)")

_PARTS_NSN_RE = re.compile(r'\b(\d{4}-\d{2}-\d{3}-\d{4})\b')
_PARTS_FIG_RE = re.compile(r'\bFIG(?:URE)?\.?\s*(\d+)\s*[:.\-]?\s*([A-Za-z][A-Za-z0-9 ,/&()\-]{2,50})')

def extract_parts(con):
    """Build a structured, cited parts index from RPSTL pages: each NSN tied to its figure
    (number + title), document, page, and vehicle. Reliable + idempotent (full rebuild). Exact
    NSN->part#/nomenclature row alignment is OCR-noisy, so it is intentionally NOT asserted here;
    every record carries its source page for citation/verification."""
    log("extracting structured parts from RPSTL pages...")
    dbdir = _db_dir(con)
    # Review finding: ocrall's heartbeat only covered the OCR batch loop, not this phase that runs
    # after it -- ocr_supervisor.py could mistake a long extract_parts() pass for a stale hang and
    # kill it, discarding the work (and run_ocr_auto.bat would still declare "OCR COMPLETE" since
    # OCR itself had already reached 0 pending by then).
    _heartbeat(dbdir, 0, 0, "extract_parts")
    if not _EXTRACT_TALLY: _tally_reset()   # same lazy-default as crawl()/ocr(): the bare `parts` subcommand calls this alone
    _write_progress(dbdir, stage="parts", current=None, done=0, total=None, extracted=dict(_EXTRACT_TALLY))
    con.execute("DELETE FROM parts WHERE confidence IS NOT NULL"); con.commit()
    try:
        rows = con.execute(
            "SELECT p.document_id, p.page_number, p.body_text, d.vehicle "
            "FROM pages_fts JOIN pages p ON p.id=pages_fts.rowid JOIN documents d ON d.id=p.document_id "
            "WHERE pages_fts MATCH ?", ('"usable on code"',)).fetchall()
    except Exception as e:
        log(f"parts: FTS query failed ({e}); run migrate/crawl first"); return 0
    batch = []; seen = set(); npages = 0
    for r in rows:
        # Bug found + fixed during review verification (pre-existing, predates this diff): this
        # loop used dict-style row access (r["document_id"], ...) against a connection that
        # returns plain tuples -- connect() sets no row_factory anywhere in this file. Reproduced
        # directly: extract_parts() raised "tuple indices must be integers or slices, not str" on
        # the first RPSTL-style page matching the FTS phrase below, which real parts-list text
        # commonly does. Switched to positional unpacking, matching every other query in this file.
        document_id, page_number, body_text, vehicle = r
        npages += 1
        bt = body_text or ""
        figs = [(m.start(), m.group(1), re.sub(r'\s+', ' ', m.group(2)).strip()) for m in _PARTS_FIG_RE.finditer(bt)]
        for m in _PARTS_NSN_RE.finditer(bt):
            nsn = m.group(1); pos = m.start(); fno = ftit = None
            for fp, fn, ft in figs:
                if fp <= pos: fno, ftit = fn, ft
                else: break
            key = (document_id, page_number, nsn)
            if key in seen: continue
            seen.add(key)
            batch.append((nsn, ftit, ftit, document_id, page_number, vehicle, fno, ftit, "page"))
        if npages % 200 == 0:
            _heartbeat(dbdir, npages, 0, "extract_parts")
    # Barcode-decoded NSNs (migration 0010 / viewer_ingest.ocr_one()'s _scan_barcode()): a machine
    # read has no character-recognition ambiguity, so these get their own confidence tag ('barcode')
    # instead of 'page' -- distinguishable, higher-trust provenance, picked up for free by every
    # existing confidence-IS-NOT-NULL consumer (features/parts_feature.py's part_lookup() etc.)
    # without those callers needing to change. Full-rebuild-safe: this scans EVERY page with a
    # captured barcode_nsn each time extract_parts() runs (same idempotent-rebuild contract as the
    # regex pass above), so a barcode row survives the DELETE at the top of this function exactly
    # like a regex-extracted 'page' row does -- it just gets regenerated from pages.barcode_nsn
    # instead of re-parsed from body_text.
    nbar = 0
    seen_barcode = set()
    try:
        brows = con.execute(
            "SELECT p.document_id, p.page_number, p.barcode_nsn, p.barcode_data, d.vehicle "
            "FROM pages p JOIN documents d ON d.id=p.document_id "
            "WHERE p.barcode_nsn IS NOT NULL AND p.barcode_nsn <> ''").fetchall()
    except sqlite3.OperationalError:
        brows = []   # pre-migration-0010 schema (barcode_nsn column doesn't exist yet)
    for document_id, page_number, bnsn, bdata, vehicle in brows:
        # Review finding: normalize through patterns.norm_nsn(), the SAME canonical-NSN helper every
        # other NSN producer in this codebase uses. barcodes.py's own NSN regex (barcodes._NSN)
        # tolerates a dashless 13-digit payload (a common Code39/128 DoD barcode convention);
        # storing that verbatim would defeat every downstream lookup that normalizes its query before
        # matching `WHERE nsn=?` (e.g. features/parts_feature.py's part_differences()), and would
        # never collide -- via the dedup key below -- with a same-part regex-extracted row that
        # already normalized to the canonical dashed form. norm_nsn() always succeeds here since
        # barcodes._NSN's match shape is a subset of patterns.NSN_RE's; `or bnsn` is a pure defensive
        # fallback, never expected to be exercised.
        nsn = norm_nsn(bnsn) or bnsn
        # Review finding: this used to share the regex pass's `seen` set (keyed on (document_id,
        # page_number, nsn)) -- when a page's RPSTL text and its barcode both encode the same NSN,
        # the higher-trust barcode row this feature exists to surface was silently dropped in favor
        # of the lower-trust 'page' row that happened to run first, exactly defeating the feature on
        # its own corroborating case. A barcode-derived row is never a duplicate of a regex-derived
        # one, so it gets its own dedup key scoped to this loop (defensive only: brows already
        # carries at most one row per (document_id, page_number), since barcode_nsn is a single
        # column on pages, so this can't actually fire today).
        key = (document_id, page_number, nsn)
        if key in seen_barcode: continue
        seen_barcode.add(key)
        label = (bdata or "").strip() or None
        batch.append((nsn, label, label, document_id, page_number, vehicle, None, None, "barcode"))
        nbar += 1
    con.executemany(
        "INSERT INTO parts(nsn, name, nomenclature, document_id, page, vehicle, fig_no, fig_title, confidence, created_at) "
        "VALUES(?,?,?,?,?,?,?,?,?,datetime('now'))", batch)
    con.commit()
    _heartbeat(dbdir, npages, 0, 0)
    # Final, authoritative counts -- extract_parts() does a full DELETE-then-rebuild every time (see
    # its own docstring), so these OVERWRITE rather than accumulate, matching that same contract:
    # the true count as of right now, not a running total across repeated calls in one process.
    _EXTRACT_TALLY["parts_page"] = len(batch) - nbar
    _EXTRACT_TALLY["parts_barcode"] = nbar
    # A few real, distinct NSNs -- not just a count -- so the in-app scan UI can link straight to
    # /part?q=<nsn> ("View parts") instead of a dead-end generic page. Capped small; this is a
    # "here's a sample of what showed up" pointer, not meant to be the parts index's own listing.
    _EXTRACT_TALLY["nsn_samples"] = list(dict.fromkeys(b[0] for b in batch))[:5]
    _write_progress(dbdir, stage="parts", current=None, done=npages, total=npages, extracted=dict(_EXTRACT_TALLY))
    log(f"parts: {npages} RPSTL pages -> {len(batch)} NSN records ({len(set(b[0] for b in batch))} distinct NSNs)"
        f"{f', {nbar} from decoded barcodes' if nbar else ''}")
    return len(batch)


# ---- schematics: page-level detection, live -----------------------------------------------------
# Two independent signals, mirroring build_schemgraph.py's own per-doc scan loop (schem_overlay.
# schem_paths() -> schemgraph.graph_from_paths(), same >=12-path has_vector gate and >=8-edge
# min-edges floor) but scoped to just the documents THIS run touched, not the whole corpus, and
# additionally catching scanned/raster schematic pages the vector path can never see at all.
_SCHEM_KEYWORDS_RE = re.compile(
    r'\b(?:SCHEMATIC|WIRING\s+DIAGRAM|CIRCUIT\s+DIAGRAM|ELECTRICAL\s+DIAGRAM|'
    r'HYDRAULIC\s+DIAGRAM|PNEUMATIC\s+DIAGRAM)\b', re.I)
_SCHEM_MIN_EDGES = 8   # same floor build_schemgraph.py uses (--min-edges default) -- filters out
                       # pages with a little incidental vector content (a ruled table, a border)
                       # that aren't actually wiring diagrams.

def detect_schematics(con, doc_id, path):
    """Per-document schematic detection -- run once a document's pages all have real text (direct
    or OCR'd), so the keyword signal has something to search. Best-effort throughout: any failure
    for one page or the whole document must never break the ingest job it's called from.
    Returns the count of (doc,page) schematic rows written (0 on any failure, or instantly when
    SCHEMATIC_SCAN is off -- this is the one extraction toggle with a real per-page PDF-reopen
    cost; see its module-level comment)."""
    if not SCHEMATIC_SCAN:
        return 0
    dbdir = _db_dir(con)
    try:
        import schem_overlay, schemgraph
    except Exception:
        schem_overlay = schemgraph = None
    rows = con.execute("SELECT page_number, body_text FROM pages WHERE document_id=? ORDER BY page_number",
                       (doc_id,)).fetchall()
    if not rows:
        return 0
    vrow = con.execute("SELECT vehicle FROM documents WHERE id=?", (doc_id,)).fetchone()
    vehicle = vrow[0] if vrow else None
    is_pdf = bool(path) and str(path).lower().endswith(".pdf") and os.path.exists(path)
    n = 0
    for page_number, body_text in rows:
        detected_via = None; has_netlist = 0; net_count = comp_count = confidence = None; caption = None
        if is_pdf and schem_overlay is not None and schemgraph is not None:
            try:
                raw = schem_overlay.schem_paths(path, page_number)
                if raw.get("has_vector"):
                    g = schemgraph.graph_from_paths(raw)
                    edges = (g.get("counts") or {}).get("edges", 0)
                    if edges >= _SCHEM_MIN_EDGES:
                        detected_via = "vector"
                        c = g.get("counts") or {}
                        net_count = c.get("nets"); comp_count = c.get("components"); confidence = g.get("confidence")
                        if dbdir:
                            try:
                                import safeguard
                                cache_dir = os.path.join(dbdir, "schemcache")
                                os.makedirs(cache_dir, exist_ok=True)
                                g["page"] = page_number
                                # same location + filename build_schemgraph.py already writes
                                # (schemgraph.cache_path()) -- /schemflow.js and Circuit Lab's
                                # reference panel pick this up with no changes of their own.
                                safeguard.atomic_write(schemgraph.cache_path(cache_dir, doc_id, page_number),
                                                       json.dumps(g))
                                has_netlist = 1
                            except Exception:
                                pass
            except Exception:
                pass
        if detected_via is None and body_text:
            m = _SCHEM_KEYWORDS_RE.search(body_text)
            if m:
                detected_via = "keyword"
                caption = re.sub(r'\s+', ' ', body_text[max(0, m.start() - 40):m.end() + 40]).strip()
        if detected_via:
            try:
                con.execute(
                    "INSERT INTO schematics(document_id,page_number,vehicle,detected_via,has_netlist,"
                    "net_count,component_count,confidence,caption) VALUES(?,?,?,?,?,?,?,?,?) "
                    "ON CONFLICT(document_id,page_number) DO UPDATE SET vehicle=excluded.vehicle, "
                    "detected_via=excluded.detected_via, has_netlist=excluded.has_netlist, "
                    "net_count=excluded.net_count, component_count=excluded.component_count, "
                    "confidence=excluded.confidence, caption=excluded.caption",
                    (doc_id, page_number, vehicle, detected_via, has_netlist, net_count, comp_count,
                     confidence, caption))
                n += 1
            except sqlite3.OperationalError:
                pass   # pre-migration-0011 schema (schematics table doesn't exist yet) -- skip, never crash ingest
    if n:
        con.commit()
    return n


def _run_schematic_stage(con):
    """Shared 4th stage for both 'run' and 'ocrall': after extract_parts() has finished, detect
    schematics ONLY on the documents this process actually touched this run (_TOUCHED_DOC_IDS --
    populated by crawl()'s discoveries and ocr()'s completions above), not the whole corpus every
    time. No-op instantly if SCHEMATIC_SCAN is off or nothing was touched. Updates
    _EXTRACT_TALLY['schematics'] and stamps its own 'schematics' progress stage, same shape every
    other stage already uses."""
    if not SCHEMATIC_SCAN or not _TOUCHED_DOC_IDS:
        return
    dbdir = _db_dir(con)
    doc_ids = sorted(_TOUCHED_DOC_IDS)
    total = len(doc_ids)
    _write_progress(dbdir, stage="schematics", current=None, done=0, total=total, extracted=dict(_EXTRACT_TALLY))
    n_schem = 0
    for i, doc_id in enumerate(doc_ids, 1):
        try:
            prow = con.execute("SELECT path FROM documents WHERE id=?", (doc_id,)).fetchone()
            if prow and prow[0]:
                n_schem += detect_schematics(con, doc_id, prow[0])
        except Exception:
            pass
        _EXTRACT_TALLY["schematics"] = n_schem
        _write_progress(dbdir, stage="schematics", current={"doc": doc_id}, done=i, total=total,
                        extracted=dict(_EXTRACT_TALLY))


_TABLES_SCHEMA = """
CREATE TABLE IF NOT EXISTS tbl(
  id INTEGER PRIMARY KEY, doc INTEGER, page INTEGER, n_rows INTEGER, n_cols INTEGER,
  spec INTEGER, units TEXT);
CREATE INDEX IF NOT EXISTS ix_tbl_doc  ON tbl(doc);
CREATE INDEX IF NOT EXISTS ix_tbl_spec ON tbl(spec);
"""   # identical to build_tables.py's own SCHEMA -- same tables.db, same table name/columns, so
      # /api/tables and anything else that already reads this sidecar picks up live-extracted rows
      # with zero changes of their own (deliberately NOT including tbl_done -- that resumability
      # marker is build_tables.py's own whole-corpus-scan bookkeeping; this wiring is scoped to just
      # the documents this run touched instead, same as SCHEMATIC_SCAN's own dedicated cache file).

def extract_tables_for_doc(con, doc_id, path):
    """Per-document table extraction -- tables.extract_page() (PyMuPDF find_tables()) over every
    page, mirrored call-for-call from build_tables.py's own per-doc loop, writing into the SAME
    tables.db `tbl` schema that batch tool already defines. Best-effort throughout: any failure for
    one page or the whole document must never break the ingest job it's called from.
    Returns the count of tables written (0 on any failure, or instantly when TABLES_SCAN is off)."""
    if not TABLES_SCAN:
        return 0
    if not path or not str(path).lower().endswith(".pdf") or not os.path.exists(path):
        return 0
    try:
        import tables as _tables
        if not _tables.available():
            return 0
    except Exception:
        return 0
    dbdir = _db_dir(con)
    if not dbdir:
        return 0
    npages = con.execute("SELECT COUNT(*) FROM pages WHERE document_id=?", (doc_id,)).fetchone()[0]
    if not npages:
        return 0
    n = 0
    try:
        tdb = sqlite3.connect(os.path.join(dbdir, "tables.db"))
        tdb.executescript(_TABLES_SCHEMA)
        tdb.execute("DELETE FROM tbl WHERE doc=?", (doc_id,))   # idempotent-rebuild, same contract extract_parts() already uses
        for pg in range(1, npages + 1):
            try:
                for t in _tables.extract_page(path, pg):
                    tdb.execute("INSERT INTO tbl(doc,page,n_rows,n_cols,spec,units) VALUES(?,?,?,?,?,?)",
                               (doc_id, pg, t["n_rows"], t["n_cols"], 1 if t["spec"] else 0, ",".join(t["units"])))
                    n += 1
            except Exception:
                continue
        if n:
            tdb.commit()
        tdb.close()
    except Exception:
        return 0
    return n

def _run_tables_stage(con):
    """Shared stage for both 'run' and 'ocrall', alongside _run_schematic_stage() -- same
    _TOUCHED_DOC_IDS scoping, same no-op-if-off/nothing-touched contract, same progress-stage
    shape. Kept as its own independently-toggleable stage (not folded into the schematics one)
    so each real per-page-PDF-reopen cost stays separately controllable."""
    if not TABLES_SCAN or not _TOUCHED_DOC_IDS:
        return
    dbdir = _db_dir(con)
    doc_ids = sorted(_TOUCHED_DOC_IDS)
    total = len(doc_ids)
    _write_progress(dbdir, stage="tables", current=None, done=0, total=total, extracted=dict(_EXTRACT_TALLY))
    n_tables = 0
    for i, doc_id in enumerate(doc_ids, 1):
        try:
            prow = con.execute("SELECT path FROM documents WHERE id=?", (doc_id,)).fetchone()
            if prow and prow[0]:
                n_tables += extract_tables_for_doc(con, doc_id, prow[0])
        except Exception:
            pass
        _EXTRACT_TALLY["tables"] = n_tables
        _write_progress(dbdir, stage="tables", current={"doc": doc_id}, done=i, total=total,
                        extracted=dict(_EXTRACT_TALLY))


_HW_SEED = [
    # (size, series, major_in, major_mm, tpi_or_pitch, tap_drill, torque_ref_lbft)  -- public-domain facts
    ("1/4-20 UNC","UNC",0.250,None,"20","#7 (.201)","8"),
    ("5/16-18 UNC","UNC",0.3125,None,"18","F (.257)","17"),
    ("3/8-16 UNC","UNC",0.375,None,"16","5/16 (.3125)","30"),
    ("7/16-14 UNC","UNC",0.4375,None,"14","U (.368)","50"),
    ("1/2-13 UNC","UNC",0.500,None,"13","27/64 (.4219)","75"),
    ("9/16-12 UNC","UNC",0.5625,None,"12","31/64 (.4844)","110"),
    ("5/8-11 UNC","UNC",0.625,None,"11","17/32 (.5312)","150"),
    ("3/4-10 UNC","UNC",0.750,None,"10","21/32 (.6562)","270"),
    ("7/8-9 UNC","UNC",0.875,None,"9","49/64 (.7656)","430"),
    ("1-8 UNC","UNC",1.000,None,"8","7/8 (.875)","640"),
    ("1/4-28 UNF","UNF",0.250,None,"28","#3 (.213)","10"),
    ("5/16-24 UNF","UNF",0.3125,None,"24","I (.272)","19"),
    ("3/8-24 UNF","UNF",0.375,None,"24","Q (.332)","35"),
    ("1/2-20 UNF","UNF",0.500,None,"20","29/64 (.4531)","85"),
    ("5/8-18 UNF","UNF",0.625,None,"18","37/64 (.5781)","170"),
    ("3/4-16 UNF","UNF",0.750,None,"16","11/16 (.6875)","300"),
    ("M6x1.0","metric",None,6.0,"1.0","5.0 mm","7"),
    ("M8x1.25","metric",None,8.0,"1.25","6.8 mm","18"),
    ("M10x1.5","metric",None,10.0,"1.5","8.5 mm","37"),
    ("M12x1.75","metric",None,12.0,"1.75","10.2 mm","64"),
    ("M14x2.0","metric",None,14.0,"2.0","12.0 mm","100"),
    ("M16x2.0","metric",None,16.0,"2.0","14.0 mm","160"),
]

def _iter_tabular(path):
    """Yield dict rows from a .csv or .xlsx/.xlsm file (streaming, for large extracts)."""
    ext = os.path.splitext(path)[1].lower()
    if ext in (".xlsx", ".xlsm"):
        try:
            from openpyxl import load_workbook
        except Exception:
            log("enrich: openpyxl not installed -- run `pip install openpyxl`, or save the file as CSV and use --gsa file.csv")
            return
        ws = load_workbook(path, read_only=True, data_only=True).active
        it = ws.iter_rows(values_only=True)
        try:
            header = [("" if c is None else str(c)).strip() for c in next(it)]
        except StopIteration:
            return
        for r in it:
            yield {header[i]: ("" if v is None else str(v)) for i, v in enumerate(r) if i < len(header)}
    else:
        import csv as _csv
        with open(path, encoding="utf-8", errors="replace", newline="") as fh:
            for row in _csv.DictReader(fh):
                yield row

def rollback(con, confirm=False):
    """Reverse the additive reference enrichment + structured-parts extraction (R1 rollback).
    Drops the external reference tables and clears the extracted parts; the index returns to its
    pre-enrichment search behavior. Additive columns/migrations remain (harmless, backwards-compatible).
    OCR-filled page text is NOT touched (it only ADDED text to blank pages -- R6)."""
    n_ref = n_log = n_hw = n_parts = 0
    try: n_ref = con.execute("SELECT COUNT(*) FROM ref_nsn").fetchone()[0]
    except sqlite3.OperationalError: pass
    try: n_log = con.execute("SELECT COUNT(*) FROM ref_nsn_log").fetchone()[0]
    except sqlite3.OperationalError: pass
    try: n_hw = con.execute("SELECT COUNT(*) FROM ref_hardware").fetchone()[0]
    except sqlite3.OperationalError: pass
    try: n_parts = con.execute("SELECT COUNT(*) FROM parts WHERE confidence IS NOT NULL").fetchone()[0]
    except sqlite3.OperationalError: pass
    log(f"rollback would remove: ref_nsn={n_ref}, ref_nsn_log={n_log}, ref_hardware={n_hw}, extracted parts={n_parts}")
    if not confirm:
        log("DRY RUN -- nothing changed. Re-run with --yes to actually roll back."); return
    for tbl in ("ref_nsn", "ref_nsn_log", "ref_hardware"):
        try: con.execute(f"DROP TABLE IF EXISTS {tbl}")
        except sqlite3.OperationalError as e: log(f"  ({tbl}: {e})")
    try: con.execute("DELETE FROM parts WHERE confidence IS NOT NULL")
    except sqlite3.OperationalError: pass
    con.commit()
    log("ROLLED BACK: enrichment + extracted parts removed. Search/sheet behavior restored to pre-enrichment.")

def enrich_flis(con, folder):
    """Ingest the DLA FLIS Reading Room table extracts (NIIN-keyed), pure-Python streaming (Windows-ok):
      V_FLIS_IDENTIFICATION (NIIN->INC) + P_H6_PICK (INC->item name); V_FLIS_PART (part#/CAGE);
      V_FLIS_MANAGEMENT (AAC + unit price); V_CHARACTERISTICS (aggregated size/thread/material);
      V_FLIS_CANCELLED_NIIN (status/replacement, kept per R6). Filters to in-index NIINs; append-only."""
    import csv as _csv
    def find(*names):
        for nm in names:
            p = os.path.join(folder, nm)
            if os.path.exists(p): return p
        return None
    nsns = set()
    for t in ("documents", "parts"):
        try:
            for (n,) in con.execute(f"SELECT DISTINCT nsn FROM {t} WHERE nsn IS NOT NULL AND nsn<>''"): nsns.add(n.strip())
        except sqlite3.OperationalError: pass
    niin2nsn = {}
    for n in nsns:
        d = re.sub(r"\D", "", n)
        if len(d) >= 13: niin2nsn[d[4:13]] = n
    if not niin2nsn:
        log("enrich_flis: no NSNs in the index to match"); return 0
    niinset = set(niin2nsn)
    log(f"enrich_flis: matching {len(niinset)} index NIINs against FLIS files in {folder} (streaming; large files take a few minutes)")
    def stream(path):
        if not path: return
        with open(path, encoding="utf-8", errors="replace", newline="") as fh:
            rd = _csv.reader(fh); next(rd, None)
            for r in rd: yield r
    inc2name = {}
    for r in (stream(find("P_H6_PICK.CSV")) or []):
        if len(r) >= 2: inc2name[r[0].strip().zfill(5)] = r[1].strip()
    name = {}; part = {}; cage = {}; aac = {}; price = {}; char = {}; cancel = {}
    niin_inc = {}; cage_company = {}; colloq = {}; alt = {}
    for r in (stream(find("V_FLIS_IDENTIFICATION.CSV")) or []):
        if len(r) >= 2 and r[0].strip() in niinset:
            inc = r[1].strip().zfill(5); niin_inc[r[0].strip()] = inc
            name.setdefault(r[0].strip(), inc2name.get(inc, ""))
    for r in (stream(find("V_FLIS_PART.CSV")) or []):
        k = r[0].strip() if r else ""
        if len(r) >= 3 and k in niinset and k not in part: part[k] = r[1].strip(); cage[k] = r[2].strip()
    for r in (stream(find("V_FLIS_MANAGEMENT.CSV")) or []):
        k = r[0].strip() if r else ""
        if len(r) >= 9 and k in niinset and k not in aac:
            aac[k] = r[3].strip(); up = re.sub(r"^0+", "", r[8].strip()) or "0"
            try: price[k] = f"{float(up):,.2f}"
            except Exception: pass
    for r in (stream(find("V_CHARACTERISTICS.CSV")) or []):
        if len(r) >= 4 and r[0].strip() in niinset: char.setdefault(r[0].strip(), []).append(f"{r[2].strip()}: {r[3].strip()}")
    for r in (stream(find("V_FLIS_CANCELLED_NIIN.CSV")) or []):
        if len(r) >= 4 and r[0].strip() in niinset:
            cancel[r[0].strip()] = f"status {r[3].strip()}" + (f", repl NIIN {r[2].strip()}" if r[2].strip() else "")
    # CAGE -> manufacturer name (who actually makes the part); only the CAGE codes we reference.
    needed_cages = set(v for v in cage.values() if v)
    if needed_cages:
        for r in (stream(find("P_CAGE.CSV")) or []):  # CAGE_CODE,CAGE_STATUS,TYPE,CAO,COMPANY,CITY,STATE,...
            c0 = r[0].strip() if r else ""
            if len(r) >= 5 and c0 in needed_cages and c0 not in cage_company:
                loc = ", ".join(x for x in [(r[5].strip() if len(r) > 5 else ""), (r[6].strip() if len(r) > 6 else "")] if x)
                cage_company[c0] = r[4].strip() + ((" (" + loc + ")") if loc else "")
    # INC -> colloquial / common name (so 'battery', 'wrench', etc. are recognisable).
    needed_incs = set(niin_inc.values())
    if needed_incs:
        for r in (stream(find("V_COLLOQUIAL_NAME.CSV")) or []):  # INC,RELATED_INC,COLLOQUIAL_NAME
            if len(r) >= 3 and r[0].strip().zfill(5) in needed_incs and r[2].strip():
                colloq.setdefault(r[0].strip().zfill(5), r[2].strip())
    # FLIS standardization -> interchangeable / related NSNs (kept in alt_parts).
    for r in (stream(find("V_FLIS_STANDARDIZATION.CSV")) or []):  # NIIN,RELATED_NSN,ISC,...
        if len(r) >= 2 and r[0].strip() in niinset and r[1].strip():
            alt.setdefault(r[0].strip(), []).append(r[1].strip())
    SRC = "PUB LOG (DLA FLIS Reading Room, publicly releasable)"
    URL = "https://www.dla.mil/Information-Operations/FLIS-Data-Electronic-Reading-Room/"
    n = 0
    for niin, nsn in niin2nsn.items():
        itnm = name.get(niin, ""); pno = part.get(niin, ""); cg = cage.get(niin, "")
        ac = aac.get(niin, ""); pr = price.get(niin, ""); ch = "; ".join(char.get(niin, [])[:12])
        subs = cancel.get(niin, "")
        mfr = cage_company.get(cg, "") if cg else ""
        coll = colloq.get(niin_inc.get(niin, ""), "")
        altp = "; ".join(list(dict.fromkeys(alt.get(niin, [])))[:20])
        desc = "; ".join(x for x in [
            ("Made by " + mfr + " (CAGE " + cg + ")") if mfr else "",
            ("Also called: " + coll) if coll else "",
            ("Unit price $" + pr) if pr else ""] if x)
        if not any([itnm, pno, cg, ac, ch, subs, altp]): continue
        con.execute("INSERT INTO ref_nsn_log(nsn,item_name,description,part_no,cagec,characteristics,aac,substitutes,alt_parts,source,source_url,fetched_at) VALUES(?,?,?,?,?,?,?,?,?,?,?,datetime('now'))",
                    (nsn, itnm, desc, pno, cg, ch, ac, subs, altp, SRC, URL))
        con.execute("INSERT INTO ref_nsn(nsn,item_name,description,part_no,cagec,characteristics,aac,substitutes,alt_parts,source,source_url,fetched_at,official) VALUES(?,?,?,?,?,?,?,?,?,?,?,datetime('now'),1) "
                    "ON CONFLICT(nsn) DO UPDATE SET "
                    "item_name=COALESCE(NULLIF(excluded.item_name,''),item_name),description=COALESCE(NULLIF(excluded.description,''),description),"
                    "part_no=COALESCE(NULLIF(excluded.part_no,''),part_no),cagec=COALESCE(NULLIF(excluded.cagec,''),cagec),"
                    "characteristics=COALESCE(NULLIF(excluded.characteristics,''),characteristics),aac=COALESCE(NULLIF(excluded.aac,''),aac),"
                    "substitutes=COALESCE(NULLIF(excluded.substitutes,''),substitutes),alt_parts=COALESCE(NULLIF(excluded.alt_parts,''),alt_parts),source=excluded.source,source_url=excluded.source_url,fetched_at=excluded.fetched_at",
                    (nsn, itnm, desc, pno, cg, ch, ac, subs, altp, SRC, URL))
        n += 1
    con.commit()
    log(f"enrich_flis: enriched {n} NSNs from the FLIS Reading Room catalog (append-only, R6)")
    return n

def enrich(con, gsa_csv=None, publog_csv=None, publog_dir=None):
    """One-time ONLINE -> OFFLINE reference enrichment, kept separate and fully cited. The running
    engine never goes online; this is a one-shot filler you run by hand (ideally on a connected
    machine, then copy the DB back).
    (1) Public-domain standard-hardware dimensions seed (FED-STD-H28). (2) Optional official GSA NSN
    Extract CSV, ingested ONLY for NSNs already in the index. The engine stays offline after."""
    SRC = "FED-STD-H28 (public domain) + standard fastener references"
    URL = "https://everyspec.com/FED-STD/FED-STD-H28B_56183/"
    con.execute("DELETE FROM ref_hardware")
    con.executemany(
        "INSERT INTO ref_hardware(size,series,major_in,major_mm,tpi_or_pitch,tap_drill,torque_ref_lbft,source,source_url,fetched_at) "
        "VALUES(?,?,?,?,?,?,?,?,?,datetime('now'))",
        [(s, se, mi, mm, tp, td, tq, SRC, URL) for (s, se, mi, mm, tp, td, tq) in _HW_SEED])
    con.commit()
    log(f"enrich: loaded {len(_HW_SEED)} public-domain hardware reference rows (torque = general reference; TM value governs)")
    if not gsa_csv and not publog_csv:
        log("enrich: hardware reference loaded; no --gsa/--publog extract given, so NSN ingest skipped. Done."); return
    idx = set()
    for (n,) in con.execute("SELECT DISTINCT nsn FROM documents WHERE nsn IS NOT NULL AND nsn<>''"): idx.add(n.strip())
    for (n,) in con.execute("SELECT DISTINCT nsn FROM parts WHERE nsn IS NOT NULL AND nsn<>''"): idx.add(n.strip())
    log(f"enrich: {len(idx)} distinct NSNs in the index to match against the extract(s)")

    def _nrm(raw):
        d = re.sub(r"\D", "", raw or "")
        return f"{d[0:4]}-{d[4:6]}-{d[6:9]}-{d[9:13]}" if len(d) >= 13 else (raw or "").strip()
    def _g(row, *names):
        for nm in names:
            v = row.get(nm)
            if v not in (None, ""): return str(v).strip()
        return ""

    if gsa_csv:
        if not os.path.exists(gsa_csv):
            log(f"enrich: GSA file not found at {gsa_csv}")
        else:
            SRCN = "GSA National Stock Number Extract (data.gov, CC0)"
            URLN = "https://catalog.data.gov/dataset/national-stock-number-extract"
            added = 0
            for row in _iter_tabular(gsa_csv):
                nsn = _nrm(_g(row, "NSN","nsn","NationalStockNumber","National Stock Number"))
                if not nsn or nsn not in idx: continue
                name = _g(row, "ItemName","Item Name","ITEM_NAME","Item Description","Description")
                desc = _g(row, "LongDescription","Description","Item Description")
                price = _g(row, "Price","UnitPrice","GSA Price","Unit Price")
                con.execute("INSERT INTO ref_nsn_log(nsn,item_name,description,gsa_price,source,source_url,fetched_at) VALUES(?,?,?,?,?,?,datetime('now'))", (nsn,name,desc,price,SRCN,URLN))
                con.execute("INSERT INTO ref_nsn(nsn,item_name,description,gsa_price,source,source_url,fetched_at,official) VALUES(?,?,?,?,?,?,datetime('now'),1) "
                            "ON CONFLICT(nsn) DO UPDATE SET item_name=excluded.item_name,description=excluded.description,gsa_price=excluded.gsa_price,source=excluded.source,source_url=excluded.source_url,fetched_at=excluded.fetched_at", (nsn,name,desc,price,SRCN,URLN))
                added += 1
            con.commit(); log(f"enrich: GSA extract -> {added} NSN versions appended (R6: history retained)")

    URLP = "https://www.dla.mil/Information-Operations/FLIS-Data-Electronic-Reading-Room/"
    def _ingest_publog(path):
        """Ingest one PUB LOG Reading Room CSV/XLSX (Identification / Reference / Characteristics /
        Management / History / H-Series). Flexible columns; merges fields per NSN without clobbering
        values another file already supplied. Append-only log (R6)."""
        if not os.path.exists(path):
            log(f"enrich: PUB LOG file not found at {path}"); return 0
        src = "PUB LOG " + os.path.basename(path) + " (DLA FLIS Reading Room, publicly releasable)"
        n = 0
        for row in _iter_tabular(path):
            nsn = _nrm(_g(row, "NSN","nsn","NATIONAL_STOCK_NUMBER","National Stock Number"))
            if not nsn:
                fsc = _g(row, "FSC","FSC_CD","FSC_CODE"); niin = _g(row, "NIIN","NIIN_CD")
                if fsc and niin: nsn = _nrm(fsc + niin)
            if not nsn or nsn not in idx: continue
            name = _g(row, "ITEM_NAME","Item_Name","ItemName","Item Name","INC_ITEM_NAME","NAME")
            part = _g(row, "PART_NUMBER","PARTNBR","PART_NBR","Part_Number","PartNumber","REFERENCE_NUMBER","REF_NUM","Reference Number")
            cage = _g(row, "CAGE","CAGEC","CAGE_CD","Cage")
            char = _g(row, "CHARACTERISTICS","CLEAR_TEXT_REPLY","Clear_Text_Reply","REQUIREMENTS_STATEMENT","Requirements_Statement","REPLY")
            aac  = _g(row, "AAC","ACQUISITION_ADVICE_CODE","Acquisition_Advice_Code")
            subs = _g(row, "SUBSTITUTES","I_AND_S","INTERCHANGEABILITY","INTERCHANGEABLE_NSN","Interchangeable")
            con.execute("INSERT INTO ref_nsn_log(nsn,item_name,part_no,cagec,characteristics,aac,substitutes,source,source_url,fetched_at) VALUES(?,?,?,?,?,?,?,?,?,datetime('now'))",
                        (nsn,name,part,cage,char,aac,subs,src,URLP))
            con.execute("INSERT INTO ref_nsn(nsn,item_name,part_no,cagec,characteristics,aac,substitutes,source,source_url,fetched_at,official) VALUES(?,?,?,?,?,?,?,?,?,datetime('now'),1) "
                        "ON CONFLICT(nsn) DO UPDATE SET "
                        "item_name=COALESCE(NULLIF(excluded.item_name,''),item_name),"
                        "part_no=COALESCE(NULLIF(excluded.part_no,''),part_no),"
                        "cagec=COALESCE(NULLIF(excluded.cagec,''),cagec),"
                        "characteristics=COALESCE(NULLIF(excluded.characteristics,''),characteristics),"
                        "aac=COALESCE(NULLIF(excluded.aac,''),aac),"
                        "substitutes=COALESCE(NULLIF(excluded.substitutes,''),substitutes),"
                        "source=excluded.source,source_url=excluded.source_url,fetched_at=excluded.fetched_at",
                        (nsn,name,part,cage,char,aac,subs,src,URLP))
            n += 1
        con.commit(); log(f"enrich: PUB LOG {os.path.basename(path)} -> {n} NSN rows merged (R6 append + non-clobbering)")
        return n
    if publog_csv:
        _ingest_publog(publog_csv)
    if publog_dir:
        flis = any(os.path.exists(os.path.join(publog_dir, x)) for x in ("V_FLIS_PART.CSV", "V_FLIS_IDENTIFICATION.CSV", "V_CHARACTERISTICS.CSV"))
        if flis:
            enrich_flis(con, publog_dir)   # DLA FLIS Reading Room table extracts (NIIN-keyed)
        else:
            import glob
            files = sorted(set(glob.glob(os.path.join(publog_dir, "*.csv")) + glob.glob(os.path.join(publog_dir, "*.CSV")) + glob.glob(os.path.join(publog_dir, "*.xlsx"))))
            if not files: log(f"enrich: no .csv/.xlsx files found in {publog_dir}")
            for fp in files: _ingest_publog(fp)

def main():
    here = os.path.dirname(os.path.abspath(__file__))
    ap = argparse.ArgumentParser()
    ap.add_argument("cmd", choices=["migrate","crawl","ocr","ocrall","prefilter","prioritize","parts","enrich","rollback","run","status","search","cleanup","prune"])
    ap.add_argument("query", nargs="?", default="")
    ap.add_argument("--root", default=os.environ.get("VIEWER_ROOT",""))
    ap.add_argument("--db", default=os.environ.get("VIEWER_DB", os.path.join(here,"..","index","viewer.db")))
    ap.add_argument("--limit", type=int, default=200)
    ap.add_argument("--ocr-limit", type=int, default=100000)
    # --workers/--dpi/--gpu default to sentinels (None / "auto"), NOT hardcoded numbers -- resolved
    # below, after parse_args(), against sysprobe.py's build_profile() (RAM-headroom-aware, GPU-aware,
    # laptop/battery-aware ocr_workers/ocr_dpi/use_gpu; cached to index/hardware_profile.json) so an
    # operator who doesn't pass these gets THAT machine-tuned plan instead of a blind
    # os.cpu_count()/200/CPU-only guess. An explicit --workers/--dpi/--gpu on the command line always
    # wins -- the sentinel is only replaced when the flag is truly absent -- so launchers that already
    # pass explicit values (run_ocr.bat's --workers, run_ocr_gpu.bat's --gpu/--workers/--dpi,
    # run_ocr_auto.bat's sysprobe-derived --workers/--dpi/[--gpu]) are unaffected.
    ap.add_argument("--workers", type=int, default=None,
                     help="OCR worker thread count (default: auto-tuned to this machine via sysprobe.py)")
    ap.add_argument("--max-files", type=int, default=0)
    ap.add_argument("--max-seconds", type=int, default=0)
    ap.add_argument("--gpu", nargs="?", const="on", default="auto", choices=["auto", "on", "off"],
                     help="use GPU (CUDA) for OCR: 'on' forces it, 'off' forces CPU, 'auto' (default) "
                          "follows sysprobe.py's hardware probe. Bare --gpu (no value) means 'on', same "
                          "as the old on/off flag.")
    ap.add_argument("--dpi", type=int, default=None,
                     help="OCR render DPI (default: auto-tuned to this machine via sysprobe.py)")
    ap.add_argument("--adaptive", action="store_true", help="adaptive DPI: render sparse pages lower (faster). Off by default so accuracy is unchanged.")
    ap.add_argument("--gsa", default="", help="path to the official GSA NSN Extract CSV/XLSX (for `enrich`)")
    ap.add_argument("--publog", default="", help="path to a single PUB LOG (DLA) CSV/XLSX (for `enrich`)")
    ap.add_argument("--publog-dir", dest="publog_dir", default="", help="folder of extracted PUB LOG Reading Room CSVs (Identification/Reference/Characteristics/Management/History/...)")
    ap.add_argument("--yes", action="store_true", help="confirm a destructive action (e.g. `rollback`, `prune`)")
    ap.add_argument("--missing-threshold", type=float, default=0.5, help="`prune` abort threshold: max fraction of indexed documents allowed to be missing before refusing (default 0.5 = 50%%)")
    args = ap.parse_args()
    global OCR_DPI, USE_CUDA, ADAPTIVE_DPI
    if getattr(args, "adaptive", False): ADAPTIVE_DPI = True

    # Resolve the --workers/--dpi/--gpu sentinels against sysprobe.py's hardware profile -- ONLY for
    # the subcommands that actually run OCR (ocr/ocrall/run); status/search/crawl/prune/etc. never
    # touch OCR_DPI/USE_CUDA/args.workers, so they never pay the probe cost. Fail-open, same precedent
    # as --adaptive above / the rest of this codebase's optional-integration try/excepts: if sysprobe
    # raises for ANY reason (import failure, a corrupt/partial index/hardware_profile.json, etc.), or
    # doesn't have every key we need, fall straight through to today's EXACT prior defaults
    # (os.cpu_count(), 200, False) untouched below.
    prof = None
    if args.cmd in ("ocr", "ocrall", "run") and (args.workers is None or args.dpi is None or args.gpu == "auto"):
        try:
            import sysprobe
            prof = sysprobe.load_or_build()
            if not isinstance(prof, dict): prof = None
        except Exception:
            prof = None
    if args.workers is None:
        args.workers = int(prof["ocr_workers"]) if prof and prof.get("ocr_workers") else max(1, (os.cpu_count() or 2))
    if args.dpi is None:
        args.dpi = int(prof["ocr_dpi"]) if prof and prof.get("ocr_dpi") else 200
    if args.gpu == "auto":
        args.gpu = bool(prof.get("use_gpu")) if prof else False
    else:
        args.gpu = (args.gpu == "on")

    OCR_DPI = args.dpi; USE_CUDA = args.gpu
    os.makedirs(os.path.dirname(os.path.abspath(args.db)), exist_ok=True)
    con = connect(args.db); migrate(con, os.path.join(here,"migrations"), db_path=args.db)
    # One extraction tally per subprocess invocation (see _tally_reset()'s docstring) -- every
    # subcommand that can write progress starts from a clean slate, not whatever a PRIOR process
    # happened to leave in the in-memory global (each CLI invocation is its own process anyway, so
    # this mostly guards direct/repeated in-process callers, e.g. tests).
    _tally_reset()
    if args.cmd == "migrate": pass
    elif args.cmd == "crawl": crawl(con, args.root, args.max_files, args.max_seconds)
    elif args.cmd == "ocr": ocr(con, args.limit, args.workers)
    elif args.cmd == "prioritize": prioritize(con)
    elif args.cmd == "parts": extract_parts(con)
    elif args.cmd == "enrich": enrich(con, args.gsa or None, args.publog or None, args.publog_dir or None)
    elif args.cmd == "rollback": rollback(con, args.yes)
    elif args.cmd == "prefilter": prefilter(con, args.limit if args.limit and args.limit>200 else 100000)
    elif args.cmd == "ocrall":
        prioritize(con)
        while ocr(con, args.limit, args.workers) > 0: pass
        extract_parts(con)   # refresh the structured parts index after OCR adds pages
        _run_schematic_stage(con)   # 4th stage: schematic detection on whatever got OCR'd this run
        _run_tables_stage(con)      # 5th stage: table extraction, same scoping
        _write_progress(_db_dir(con), stage="done", current=None, extracted=dict(_EXTRACT_TALLY))
    elif args.cmd == "run":
        # the in-app "Add documents" scan/OCR job (features/ingest_feature.py's ingest_start())
        # launches exactly this subcommand -- crawl/ocr()/extract_parts()/_run_schematic_stage()/
        # _run_tables_stage() each stamp their own stage into ingest_progress.json as they go (see
        # _write_progress() calls inside each), so the only thing left to mark here is the final
        # "done" once every stage has actually finished, for the polling UI to stop showing a
        # stage and show a completion state instead.
        crawl(con, args.root)
        while ocr(con, 200, args.workers) > 0 and args.ocr_limit > 0: args.ocr_limit -= 200
        extract_parts(con)
        _run_schematic_stage(con)   # 4th stage: schematic detection, scoped to just this run's documents
        _run_tables_stage(con)      # 5th stage: table extraction, same scoping
        _write_progress(_db_dir(con), stage="done", current=None, extracted=dict(_EXTRACT_TALLY))
    elif args.cmd == "status": status(con)
    elif args.cmd == "search": search(con, args.query)
    elif args.cmd == "cleanup": cleanup(con)
    elif args.cmd == "prune": prune(con, args.root, confirm=args.yes, missing_threshold=args.missing_threshold,
                                     index_dir=os.path.dirname(os.path.abspath(args.db)))
    con.close()

if __name__ == "__main__": main()
